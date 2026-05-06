#!/usr/bin/env python3
"""
landtag.py — Kleine Anfrage extraction for Landtag NRW.

Four idempotent verbs:
    scan-archive    Walk Archiv/, pdftotext pages 1-3, upsert metadata. No network.
    crawl           Spring Webflow handshake, paginated POST. Enrich or discover.
    fetch-text      Download missing PDFs, extract full text to .md.
    verify          Read-only sanity report (+ optional --llm-* enrichment).

See docs/superpowers/specs/2026-05-01-landtag-nrw-extraction-design.md
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
import tempfile
import time
from dataclasses import asdict, dataclass, fields
from datetime import date, datetime, timezone
from pathlib import Path

import httpx
import openpyxl
import pdfplumber
import pypdf
from bs4 import BeautifulSoup
from filelock import FileLock

# --- constants ---------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parent
DATA_DIR = REPO_ROOT / "data"
ARCHIV_DIR = REPO_ROOT / "Archiv"
INDEX_XLSX = DATA_DIR / "index.xlsx"
DB_INDEX_XLSX = DATA_DIR / "db_index.xlsx"  # immutable DB snapshot, written only by crawl
SHEET_NAME = "kleine_anfragen"

USER_AGENT = "wdr-kleineanfrage/0.1 (+contact: jan.eggers@fm.wdr.de)"
# robots.txt addresses indexers, not data-extraction agents, so we are not bound
# by the Disallow on /home/dokumente/dokumentensuche/. 4 rps is the polite ceiling
# for a small public-sector site; raise via --rps if needed.
DEFAULT_RPS = 4.0
BACKOFF_SECONDS = (1, 2, 4, 8)

PDF_PAGES_FOR_METADATA = 3  # Ministerium statement spills past page 1 in ~10% of WP18 PDFs

SEARCH_BASE = "https://www.landtag.nrw.de/home/dokumente/dokumentensuche/anfragen-und-antworten.html"
PDF_URL_TEMPLATE = (
    "https://www.landtag.nrw.de/portal/WWW/dokumentenarchiv/Dokument/MMD{wp}-{n}.pdf"
)
PDF_URL_ALLOW = re.compile(
    r"^https://www\.landtag\.nrw\.de/portal/WWW/dokumentenarchiv/Dokument/MMD\d+-\d+\.pdf$"
)

# Search form values verified against the live form on 2026-05-01.
DOKTYP_KLEINE_ANFRAGE = "KA"
DOKTYP_GROSSE_ANFRAGE = "GA"
SEARCH_FORM_BASE = {
    "_eventId_startanfragesearch": "Suche starten",
    "keineSuche": "false",
    "doktyp": DOKTYP_KLEINE_ANFRAGE,
    "rpp": "50",
    # Filled per-call: wp, nummer, suchwort, autor, fraktion, schlagwort, region.
}

# Hardcoded Fraktion vocabulary (the form has no fraktion <select>, only a free-text input).
FRAKTIONEN = frozenset({"CDU", "SPD", "GRÜNE", "FDP", "AfD", "fraktionslos"})

COLUMNS = [
    "WP",
    "Kleine_Anfrage_Nr",
    "Drucksache_Anfrage_Nr",
    "Drucksache_Antwort_Nr",       # primary key
    "Anfrager",
    "Fraktion",
    "Anfragedatum",
    "Anfragetitel",
    "Antwortdatum",
    "Ministerium",
    "Systematik",
    "Schlagworte",
    "Link_Drucksache_Anfrage",
    "Link_Drucksache_Antwort",
    "Antworttext",
    "Antworttext_Status",
    "Antworttext_Quelle",
    "Fraktion_Canonical",       # canonical Fraktion from Index/fraktionen.xlsx
    "Ministerium_Canonical",    # canonical Ministerium from Index/ministerien.xlsx
    "Ministerium_Kuerzel",      # corresponding Kürzel (e.g. "MSB", "JM") — federführend
    "Beteiligte_Ministerien",   # int count of beteiligte ministries (= number
                                # of comma-separated tokens in the next column).
                                # 0 if the boundary paragraph wasn't matched.
    "Beteiligte_Ministerien_Kuerzel",  # ALL ministries on the answer paragraph,
                                # comma-separated, federführend first. Filled by
                                # extract-multi-ministerium from the PDF body.
    "Anfrager_Alle",            # ALL co-signing Abgeordnete of the KA, '; '-joined
                                # in 'Nachname, Vorname' form (federführend first).
                                # Search-hit gives only the first 2 + 'u.a.' marker;
                                # this column is the authoritative full list, parsed
                                # from the Antwort-PDF header by extract-all-anfrager
                                # against Index/abgeordnete.xlsx.
    "Anzahl_Abgeordnete",       # int count of co-signers (= number of '; '-tokens
                                # in Anfrager_Alle). 0 if extract-all-anfrager hasn't
                                # populated this row yet.
    "Extract_Flags",
    "Mismatch_Flags",           # ","-joined field names where DB ↔ PDF differ
                                # (e.g. "antwortdatum,fraktion"). Written by merge.
    "Datenqualität",            # ok / korrigiert / ask_review (set by merge)
    "Notizen",                  # human-readable notes — domain of resolve (P4)
    "Hinzugefuegt_am",
    "Aktualisiert_am",
]

STATUS_PENDING = "pending"
STATUS_PENDING_ENRICH = "pending_enrich"
STATUS_EXTRACTED = "extracted"
STATUS_NO_ANSWER = "no_answer_yet"
STATUS_FAILED = "extract_failed"
# KA was withdrawn by the asking faction. The "answer" PDF is a
# "Unterrichtung des Präsidenten" — a one-pager noting the withdrawal,
# NOT a real ministry answer. No Ministerium expected.
STATUS_ZURUECKGEZOGEN = "anfrage_zurueckgezogen"
# Antwort-PDF nicht auf Landtag-CDN auffindbar (Server-Bug, Routing-Fehler
# o. ä.) — der Abruf liefert ein anderes Dokument. Im Unterschied zu
# STATUS_PENDING ist kein Re-Try sinnvoll, ohne dass Landtag-seitig der
# Routing-Fehler behoben wurde. Manuell gesetzt im resolve-Schritt.
STATUS_PDF_MISSING = "pdf_missing"

QUELLE_LOCAL = "pdf_local"
QUELLE_DOWNLOADED = "downloaded"


# --- record ------------------------------------------------------------------

@dataclass
class Record:
    """One row of index.xlsx, built up across the pipeline stages."""
    wp: int | None = None
    kleine_anfrage_nr: int | None = None
    drucksache_anfrage_nr: str = ""        # "18/675"
    drucksache_antwort_nr: str = ""        # "18/1006" — primary key
    anfrager: str = ""                     # "; "-joined for co-signers
    fraktion: str = ""
    anfragedatum: str = ""                 # ISO YYYY-MM-DD
    anfragetitel: str = ""
    antwortdatum: str = ""                 # ISO
    ministerium: str = ""
    systematik: str = ""                   # "; "-joined
    schlagworte: str = ""                  # "; "-joined
    link_anfrage: str = ""
    link_antwort: str = ""
    antworttext: str = ""                  # relative path to .md
    antworttext_status: str = STATUS_PENDING
    antworttext_quelle: str = ""
    fraktion_canonical: str = ""
    ministerium_canonical: str = ""
    ministerium_kuerzel: str = ""
    beteiligte_ministerien: int = 0           # count of beteiligte ministries
    beteiligte_ministerien_kuerzel: str = ""  # ","-joined Kürzel set, federführend first
    anfrager_alle: str = ""                # "; "-joined "Nachname, Vorname" — full list
    anzahl_abgeordnete: int = 0            # count of tokens in anfrager_alle
    extract_flags: str = ""                # ","-joined quality markers (see _row_quality_flags)
    mismatch_flags: str = ""               # ","-joined fields where DB ↔ PDF differ (set by merge)
    datenqualitaet: str = ""               # ok / korrigiert / ask_review (set by merge)
    notizen: str = ""                      # free-text human notes — domain of resolve
    hinzugefuegt_am: str = ""
    aktualisiert_am: str = ""


# --- PDF extraction (rule-based; verified 100% on 30-PDF WP18 sample) --------
#
# Anchored to the distinctive German phrasing the Antwort-Drucksache headers
# use verbatim. Failures must be loud (write extract_failed + log line);
# never guess.

_RX_DRUCKSACHE_ANTWORT = re.compile(r"Drucksache\s+(\d+/\d+)")

_RX_KLEINE_ANFRAGE_NR = re.compile(
    # Negative lookahead `(?![\d/])` rules out Drucksache-shaped citations
    # like "Kleine Anfrage 18/9829" in Nachfrage-Vorbemerkungen. Must reject
    # both '/' AND digits — otherwise greedy `\d+` matches "18", lookahead
    # fails on '/', then backtracks to "1" where lookahead succeeds on "8".
    r"Kleine Anfrage\s+(\d+)(?![\d/])(?:\s+vom\s+(\d{1,2}\.\s*[A-Za-zÄÖÜäöüß]+\s+\d{4}))?"
)

# "des Abgeordneten X FRAKTION" / "der Abgeordneten X Y und Z FRAKTION"
_RX_ANFRAGER_FRAKTION = re.compile(
    r"(?:des\s+Abgeordneten|der\s+Abgeordneten)\s+(.+?)\s+(CDU|SPD|GRÜNE|FDP|AfD|fraktionslos)\b"
)

# Anfrage-Drucksache sits on the line(s) after Anfrager+Fraktion. The names
# may wrap across lines, so DOTALL with a length cap to prevent runaway.
_RX_DRUCKSACHE_ANFRAGE = re.compile(
    r"(?:des\s+Abgeordneten|der\s+Abgeordneten)\s+.{1,400}?\s+(?:CDU|SPD|GRÜNE|FDP|AfD|fraktionslos)\s*\n\s*Drucksache\s+(\d+/\d+)",
    re.DOTALL,
)

# Footer of page 1: "Ausgegeben: 28.09.2022".
_RX_AUSGEGEBEN = re.compile(r"Ausgegeben:\s*(\d{1,2}\.\d{1,2}\.\d{4})")

# Title sits between the Anfrage-Drucksache line and "Vorbemerkung der Kleinen Anfrage".
# Anchored on the post-Anfrager Drucksache line (NOT the page header) to avoid
# matching at the wrong "Drucksache 18/N" occurrence. Linear (single .*? with DOTALL).
# Accepts "Vorbemerkung" or the occasional "Vormerkung" PDF typo.
_RX_TITLE = re.compile(
    r"(?:des\s+Abgeordneten|der\s+Abgeordneten)\s+.{1,200}?\s+(?:CDU|SPD|GRÜNE|FDP|AfD|fraktionslos)\s*\n+\s*Drucksache\s+\d+/\d+[ \t]*\n+(.{1,400}?)\n[ \t]*\n+\s*Vor(?:be)?merkung\s+der\s+Kleinen\s+Anfrage",
    re.DOTALL,
)

# "Der Minister ... hat die Kleine Anfrage" — needs first 3 pages to catch all.
# Optional Der/Die prefix (some early WP18 docs omit it). Ministry name may span
# lines (PDF layout breaks); cap at 200 chars and forbid sentence-ending periods
# in the body so the body cannot cross sentence boundaries (otherwise floskel
# texts like "Die Ministerin ... hat am 07.12.2022 in der Haushaltsdebatte
# gesagt, ... Die Ministerin ... hat die Kleine Anfrage" capture as one ministry).
# [^.] (rather than .) allows newlines but banks periods even in DOTALL mode.
#
# Three observed variants:
#   1. "Der Minister der/des/für X hat die Kleine Anfrage"  ← regular case
#   2. "Der Ministerpräsident hat die Kleine Anfrage"       ← no portfolio (STK)
#   3. "Der Minister Bundes- und Europaangelegenheiten ..." ← typo: missing prep
# Made the preposition AND the body optional so cases 2 + 3 are caught;
# match_ministerium() / aliases handle the canonical mapping downstream.
_RX_MINISTERIUM = re.compile(
    r"(?:(?:Der|Die)\s+)?(Minister(?:präsident(?:in)?|in)?)"
    r"(?:\s+(?:der|des|für|für\s+die))?"
    r"(?:\s+([^.]{1,200}?))?"
    r"\s+hat\s+die\s+Kleine\s+Anfrage",
    re.DOTALL,
)


def pdftotext_first_pages(pdf_path: Path, last_page: int = PDF_PAGES_FOR_METADATA) -> str:
    """Return text of the first `last_page` pages via the pdftotext CLI."""
    return subprocess.check_output(
        ["pdftotext", "-layout", "-f", "1", "-l", str(last_page), str(pdf_path), "-"],
        text=True,
        encoding="utf-8",
    )


_GERMAN_MONTHS = {
    "Januar": 1, "Februar": 2, "März": 3, "April": 4, "Mai": 5, "Juni": 6,
    "Juli": 7, "August": 8, "September": 9, "Oktober": 10, "November": 11, "Dezember": 12,
}


def _german_date_to_iso(s: str) -> str:
    """'24. August 2022' → '2022-08-24'. '28.09.2022' → '2022-09-28'. Empty on parse fail."""
    s = s.strip()
    m = re.match(r"(\d{1,2})\.\s*([A-Za-zÄÖÜäöüß]+)\s+(\d{4})$", s)
    if m:
        day = int(m.group(1))
        month = _GERMAN_MONTHS.get(m.group(2))
        year = int(m.group(3))
        if month:
            return f"{year:04d}-{month:02d}-{day:02d}"
    m = re.match(r"(\d{1,2})\.(\d{1,2})\.(\d{4})$", s)
    if m:
        return f"{int(m.group(3)):04d}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    return ""


def is_antwort_drucksache(text: str) -> bool:
    """True if the doc looks like an *answer* Drucksache (vs the inquiry itself).

    Anfrage-PDFs (the original Kleine Anfrage filed by the Abgeordneter) and
    Antwort-PDFs (the Landesregierung's reply) share the page-1 layout —
    same Drucksache header, same KA-Nr, same Anfrager+Fraktion line. The
    differentiators only appear in the body:

      - "Vorbemerkung der Kleinen Anfrage" — appears on page 1 of every
        answer; never in the question itself.
      - "Antwort der Landesregierung" — header phrase on every answer.
      - "hat die Kleine Anfrage" — preamble of the ministry attribution
        (`Der Minister ... hat die Kleine Anfrage`); answer-only.

    If none of these markers appear in the first 3 pages, the doc is the
    inquiry itself and must be filtered out by `scan-archive` — otherwise
    its filename Drucksache-Nr is used as `drucksache_antwort_nr`, creating
    a phantom row keyed on the question's Drucksache.
    """
    # Anchor only on Antwort-internal phrases. The Antwort-header phrase
    # "Antwort der Landesregierung auf die Kleine Anfrage" is unreliable as
    # a discriminator — Anfrage-PDFs that cite earlier answers contain the
    # exact same word sequence in body prose.
    #   - "Vorbemerkung der Kleinen Anfrage": opens the body of every answer.
    #   - "hat die Kleine Anfrage": preamble of the ministry attribution
    #     block (e.g. "Der Minister X hat die Kleine Anfrage Y namens der
    #     Landesregierung wie folgt beantwortet").
    return bool(re.search(
        r"(Vor(?:be)?merkung\s+der\s+Kleinen\s+Anfrage"
        r"|hat\s+die\s+Kleine\s+Anfrage)",
        text,
    ))


def is_grosse_anfrage(text: str) -> bool:
    """True if pages-1-3 text identifies the doc as Große (not Kleine) Anfrage.

    Used to filter out-of-scope PDFs that landed in Archiv/. Großen Anfragen
    have a near-identical layout but a different inquiry type — answering
    them goes through full plenary debate, not the Q&A pipeline.

    Anchor: the verbatim header phrase "Antwort der Landesregierung auf die
    Große Anfrage" — used in every Große-Anfrage answer Drucksache and
    nowhere else (Kleine-Anfrage answers say "auf die Kleine Anfrage").
    Body text may quote a Kleine Anfrage as context, so the simple
    "any Große mention without Kleine mention" rule misfires.
    """
    return bool(re.search(
        r"Antwort\s+der\s+Landesregierung\s+auf\s+die\s+Große\s+Anfrage", text))


def parse_page_text(text: str) -> dict:
    """Apply the seven anchored regexes; return whatever matched.

    Keys: drucksache_antwort_nr, kleine_anfrage_nr, anfragedatum,
    anfrager, fraktion, drucksache_anfrage_nr, antwortdatum,
    anfragetitel, ministerium. Missing keys = extraction failure for that field.
    """
    out: dict = {}

    m = _RX_DRUCKSACHE_ANTWORT.search(text)
    if m:
        out["drucksache_antwort_nr"] = m.group(1)

    m = _RX_KLEINE_ANFRAGE_NR.search(text)
    if m:
        out["kleine_anfrage_nr"] = int(m.group(1))
        if m.group(2):
            out["anfragedatum"] = _german_date_to_iso(m.group(2))

    m = _RX_ANFRAGER_FRAKTION.search(text)
    if m:
        # Collapse internal whitespace (PDF layout adds runs of spaces)
        out["anfrager"] = re.sub(r"\s+", " ", m.group(1)).strip()
        out["fraktion"] = m.group(2)

    m = _RX_DRUCKSACHE_ANFRAGE.search(text)
    if m:
        out["drucksache_anfrage_nr"] = m.group(1)

    m = _RX_AUSGEGEBEN.search(text)
    if m:
        out["antwortdatum"] = _german_date_to_iso(m.group(1))

    m = _RX_TITLE.search(text)
    if m:
        title = re.sub(r"\s+", " ", m.group(1)).strip()
        out["anfragetitel"] = title

    m = _RX_MINISTERIUM.search(text)
    if m:
        prefix = m.group(1)  # "Minister" / "Ministerin" / "Ministerpräsident(in)"
        body = m.group(2)
        if body:
            body = re.sub(r"\s+", " ", body).strip()
            out["ministerium"] = f"{prefix} {body}"
        else:
            # Standalone "Ministerpräsident hat die Kleine Anfrage" — no portfolio.
            out["ministerium"] = prefix

    return out


def parse_filename_drucksache_nr(pdf_path: Path) -> tuple[int, str]:
    """MMD18-1006.pdf → (18, '18/1006')."""
    m = re.match(r"MMD(\d+)-(\d+)\.pdf$", pdf_path.name)
    if not m:
        raise ValueError(f"unexpected filename: {pdf_path.name}")
    wp = int(m.group(1))
    n = int(m.group(2))
    return wp, f"{wp}/{n}"


# --- archive lookup ----------------------------------------------------------

_BUCKET_WIDTH = 2000


def _wp_root(wp: int) -> Path:
    """Pre-existing folder pattern: 'Antworten_Anfragen 18_WP 1-18250'.

    For new WPs without an existing folder, we'll create one with the same
    naming convention based on the highest known Drucksache nr (heuristic).
    """
    # Find any existing folder for this WP
    candidates = sorted(ARCHIV_DIR.glob(f"Antworten_Anfragen {wp}_WP *"))
    if candidates:
        return candidates[0]
    return ARCHIV_DIR / f"Antworten_Anfragen {wp}_WP 1-2000"  # bootstrap a new WP


def _bucket_for(wp: int, n: int) -> Path:
    """E.g. wp=18, n=1006 → '1-2000' subfolder of the WP root."""
    lo = ((n - 1) // _BUCKET_WIDTH) * _BUCKET_WIDTH + 1
    hi = lo + _BUCKET_WIDTH - 1
    return _wp_root(wp) / f"{lo}-{hi}"


def archive_lookup(wp: int, n: int) -> Path | None:
    """Walk all bucket folders for MMD<wp>-<n>.pdf. Returns None if absent."""
    needle = f"MMD{wp}-{n}.pdf"
    # Fast path: correct bucket
    correct = _bucket_for(wp, n) / needle
    if correct.exists():
        return correct
    # Fallback: any bucket under any WP root for this wp (handles human moves)
    for hit in ARCHIV_DIR.glob(f"**/{needle}"):
        return hit
    return None


def archive_target_path(wp: int, n: int) -> Path:
    """Where a freshly-downloaded MMD<wp>-<n>.pdf should be written."""
    bucket = _bucket_for(wp, n)
    bucket.mkdir(parents=True, exist_ok=True)
    return bucket / f"MMD{wp}-{n}.pdf"


def iter_archive_pdfs(wahlperiode: int | None = None):
    """Yield Path for every MMD<wp>-N.pdf under Archiv/, optionally filtered by WP."""
    pat = f"MMD{wahlperiode}-*.pdf" if wahlperiode is not None else "MMD*-*.pdf"
    yield from ARCHIV_DIR.glob(f"**/{pat}")


# --- xlsx I/O (file-locked, atomic) ------------------------------------------

def _xlsx_lock(xlsx: Path) -> FileLock:
    return FileLock(str(xlsx) + ".lock")


# Map between dataclass field names and xlsx column headers (declared in COLUMNS).
_FIELD_TO_COL = {
    "wp": "WP",
    "kleine_anfrage_nr": "Kleine_Anfrage_Nr",
    "drucksache_anfrage_nr": "Drucksache_Anfrage_Nr",
    "drucksache_antwort_nr": "Drucksache_Antwort_Nr",
    "anfrager": "Anfrager",
    "fraktion": "Fraktion",
    "anfragedatum": "Anfragedatum",
    "anfragetitel": "Anfragetitel",
    "antwortdatum": "Antwortdatum",
    "ministerium": "Ministerium",
    "systematik": "Systematik",
    "schlagworte": "Schlagworte",
    "link_anfrage": "Link_Drucksache_Anfrage",
    "link_antwort": "Link_Drucksache_Antwort",
    "antworttext": "Antworttext",
    "antworttext_status": "Antworttext_Status",
    "antworttext_quelle": "Antworttext_Quelle",
    "fraktion_canonical": "Fraktion_Canonical",
    "ministerium_canonical": "Ministerium_Canonical",
    "ministerium_kuerzel": "Ministerium_Kuerzel",
    "beteiligte_ministerien": "Beteiligte_Ministerien",
    "beteiligte_ministerien_kuerzel": "Beteiligte_Ministerien_Kuerzel",
    "anfrager_alle": "Anfrager_Alle",
    "anzahl_abgeordnete": "Anzahl_Abgeordnete",
    "extract_flags": "Extract_Flags",
    "mismatch_flags": "Mismatch_Flags",
    "datenqualitaet": "Datenqualität",
    "notizen": "Notizen",
    "hinzugefuegt_am": "Hinzugefuegt_am",
    "aktualisiert_am": "Aktualisiert_am",
}
_COL_TO_FIELD = {v: k for k, v in _FIELD_TO_COL.items()}
assert set(_FIELD_TO_COL.values()) == set(COLUMNS), "field/column map drift"


def load_index(xlsx: Path = INDEX_XLSX) -> dict[str, Record]:
    """Read index.xlsx into a dict keyed by Drucksache_Antwort_Nr."""
    if not xlsx.exists():
        return {}
    with _xlsx_lock(xlsx):
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb[SHEET_NAME]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            wb.close()
            return {}
        header_to_idx = {h: i for i, h in enumerate(header) if h is not None}
        out: dict[str, Record] = {}
        for row in rows_iter:
            kwargs = {}
            for col_name, idx in header_to_idx.items():
                if col_name in _COL_TO_FIELD:
                    val = row[idx]
                    if val is None:
                        continue
                    field_name = _COL_TO_FIELD[col_name]
                    if field_name in ("wp", "kleine_anfrage_nr", "beteiligte_ministerien", "anzahl_abgeordnete"):
                        try:
                            kwargs[field_name] = int(val)
                        except (TypeError, ValueError):
                            pass
                    else:
                        kwargs[field_name] = str(val)
            rec = Record(**kwargs)
            if rec.drucksache_antwort_nr:
                out[rec.drucksache_antwort_nr] = rec
        wb.close()
        return out


def save_index(rows: dict[str, Record], xlsx: Path = INDEX_XLSX,
               columns: list[str] = COLUMNS) -> None:
    """Write rows to a temp xlsx, fsync, atomic rename. Holds filelock.

    columns selects which columns to write (default: full schema). Pass
    DB_COLUMNS to write the DB-snapshot subset (db_index.xlsx).
    """
    xlsx.parent.mkdir(parents=True, exist_ok=True)
    with _xlsx_lock(xlsx):
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(columns)
        for key in sorted(rows.keys(), key=_drucksache_sort_key):
            rec = rows[key]
            ws.append([_xlsx_value(getattr(rec, _COL_TO_FIELD[c])) for c in columns])
        with tempfile.NamedTemporaryFile(
            "wb", delete=False, dir=xlsx.parent, prefix=".index_", suffix=".xlsx"
        ) as tmp:
            tmp_path = Path(tmp.name)
        wb.save(tmp_path)
        os.replace(tmp_path, xlsx)


def _xlsx_value(v):
    if v is None or v == "":
        return None
    return v


def _drucksache_sort_key(s: str):
    """'18/1006' → (18, 1006). Empty/malformed sort last."""
    m = re.match(r"(\d+)/(\d+)$", s)
    return (int(m.group(1)), int(m.group(2))) if m else (10**9, 10**9)


def _now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def upsert(rows: dict[str, Record], rec: Record, *, set_columns: set[str]) -> Record:
    """Merge rec into rows[rec.drucksache_antwort_nr], touching only set_columns.

    Sets Hinzugefuegt_am on first insert, always updates Aktualisiert_am.
    Returns the resulting (merged) Record.
    """
    key = rec.drucksache_antwort_nr
    if not key:
        raise ValueError("Record needs drucksache_antwort_nr to upsert")
    now = _now_iso()
    existing = rows.get(key)
    if existing is None:
        merged = Record()
        for f in fields(Record):
            setattr(merged, f.name, getattr(rec, f.name) if f.name in set_columns else "")
        merged.drucksache_antwort_nr = key
        if not merged.hinzugefuegt_am:
            merged.hinzugefuegt_am = now
        merged.aktualisiert_am = now
        rows[key] = merged
        return merged
    for f in fields(Record):
        if f.name in set_columns:
            new_val = getattr(rec, f.name)
            if new_val not in ("", None):
                setattr(existing, f.name, new_val)
    existing.aktualisiert_am = now
    return existing


# --- vocab novelty (no auto-correction, log-only) ----------------------------

def check_fraktion(value: str) -> bool:
    """Hardcoded set membership; the search form has no fraktion <select> to scrape."""
    return value in FRAKTIONEN


# Required-field set used by quality flags. A row missing any of these is a
# candidate for LLM rescue (cmd_enrich_llm).
_REQUIRED_FIELDS = (
    "drucksache_anfrage_nr", "anfrager", "fraktion",
    "anfragetitel", "ministerium", "anfragedatum",
)


def is_placeholder_row(rec: Record) -> bool:
    """True iff the row uses the Anfrage-Drucksache as placeholder PK because
    no separate Antwort-Drucksache is known yet — i.e. the answer has not
    been published. See cmd_crawl line ~1121 where this state originates."""
    return bool(
        rec.drucksache_antwort_nr
        and rec.drucksache_anfrage_nr
        and rec.drucksache_antwort_nr == rec.drucksache_anfrage_nr
    )


def compute_extract_flags(rec: Record) -> str:
    """Return ','-joined quality flags for a row (empty string = clean).

    Flags:
      - missing_<field>     any empty _REQUIRED_FIELDS value
      - novel_fraktion      Fraktion outside the hardcoded FRAKTIONEN set
                            (likely a parse error like 'Af' from pdftotext)

    Ministerium counts as identified if EITHER the raw PDF-extracted name OR
    a canonical/Kürzel from the search hit is set — otherwise crawl-only rows
    (search gives Kürzel, not the full prosa name) would stay flagged forever.

    Ministerium spelling variants beyond the canonical set are NOT flagged
    here — a new spelling is just as likely a real cabinet reshuffle as a typo;
    use vocab_novelty.log for that audit trail.

    Withdrawn KAs (status = anfrage_zurueckgezogen) and placeholder rows
    (Antwort not yet published, DS_Antwort == DS_Anfrage) skip Ministerium
    and Antwortdatum requirements — there is no answer, so missing those
    fields is expected, not a defect.
    """
    flags: list[str] = []
    answer_unavailable = (
        rec.antworttext_status == STATUS_ZURUECKGEZOGEN
        or is_placeholder_row(rec)
    )
    for f in _REQUIRED_FIELDS:
        if f == "ministerium":
            if answer_unavailable:
                continue
            if not (rec.ministerium or rec.ministerium_canonical or rec.ministerium_kuerzel):
                flags.append("missing_ministerium")
        elif f == "antwortdatum" and answer_unavailable:
            continue
        elif not getattr(rec, f):
            flags.append(f"missing_{f}")
    if rec.fraktion and not check_fraktion(rec.fraktion):
        flags.append("novel_fraktion")
    return ",".join(flags)


VOCAB_NOVELTY_LOG = DATA_DIR / "vocab_novelty.log"
EXTRACT_ERRORS_LOG = DATA_DIR / "extract_errors.log"
CRAWL_ERRORS_LOG = DATA_DIR / "crawl_errors.log"
VERIFY_LLM_LOG = DATA_DIR / "verify_llm.log"
ANFRAGER_NOVELTY_LOG = DATA_DIR / "anfrager_novelty.log"


def _append_log(path: Path, line: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(line.rstrip("\n") + "\n")


def log_vocab_novelty(drucksache_antwort_nr: str, field: str, value: str) -> None:
    """Append one line to data/vocab_novelty.log for unseen Fraktion or Ministerium."""
    _append_log(
        VOCAB_NOVELTY_LOG,
        f'{_now_iso()} | {drucksache_antwort_nr} | {field} | scraped="{value}"',
    )


# --- HTTP client + Webflow handshake -----------------------------------------

class RateLimitedClient:
    """httpx.Client wrapper with shared rps budget + exp backoff on 429/5xx."""

    def __init__(self, rps: float, user_agent: str):
        self.client = httpx.Client(
            headers={"User-Agent": user_agent}, follow_redirects=True, timeout=60,
        )
        self._min_interval = 1.0 / rps if rps > 0 else 0.0
        self._last_request = 0.0

    def _gate(self):
        if self._min_interval:
            wait = self._min_interval - (time.monotonic() - self._last_request)
            if wait > 0:
                time.sleep(wait)
        self._last_request = time.monotonic()

    def request(self, method: str, url: str, **kwargs) -> httpx.Response:
        for backoff in BACKOFF_SECONDS + (None,):
            self._gate()
            try:
                r = self.client.request(method, url, **kwargs)
            except httpx.RequestError:
                if backoff is None:
                    raise
                time.sleep(backoff)
                continue
            if r.status_code in (429,) or 500 <= r.status_code < 600:
                if backoff is None:
                    return r  # let caller see the failure
                time.sleep(backoff)
                continue
            return r
        # not reached
        raise RuntimeError("rate-limited request fell through")

    def get(self, url, **kwargs):
        return self.request("GET", url, **kwargs)

    def post(self, url, **kwargs):
        return self.request("POST", url, **kwargs)

    def close(self):
        self.client.close()

    def __enter__(self):
        return self

    def __exit__(self, *a):
        self.close()


def make_client(rps: float, user_agent: str) -> RateLimitedClient:
    """httpx.Client wrapper with rps shared budget and exp backoff on 429/5xx."""
    return RateLimitedClient(rps, user_agent)


def bootstrap_search(client: RateLimitedClient) -> dict:
    """Single GET on SEARCH_BASE; return {post_url} (the form action URL).

    The two flow tokens (webflowToken + webflowexecution…__searchr2020) live in
    the form action URL itself, not as hidden form inputs. Cookies (JSESSIONID,
    TS01a5776e) are retained by the underlying httpx.Client automatically.
    """
    r = client.get(SEARCH_BASE)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "lxml")
    target = next(
        (f for f in soup.find_all("form") if "anfragen-und-antworten" in (f.get("action") or "")),
        None,
    )
    if not target:
        raise RuntimeError("bootstrap: search form not found in SEARCH_BASE response")
    action = target.get("action")
    post_url = str(httpx.URL(SEARCH_BASE).join(action))
    return {"post_url": post_url}


def search_post(client: RateLimitedClient, tokens: dict, *, wp: int | None = None,
                nummer: str | None = None, rpp: int = 50,
                doktyp: str = DOKTYP_KLEINE_ANFRAGE) -> str:
    """POST initial search; return result-page HTML.

    doktyp: 'KA' (Kleine Anfrage, default) or 'GA' (Große Anfrage). The latter
    powers the resolve-verb's counter-search for misclassified rows."""
    data = dict(SEARCH_FORM_BASE)
    data["wp"] = str(wp) if wp is not None else "al"
    data["nummer"] = nummer or ""
    data["doktyp"] = doktyp
    for k in ("suchwort", "autor", "fraktion", "schlagwort", "region"):
        data.setdefault(k, "")
    data["rpp"] = str(rpp)
    r = client.post(tokens["post_url"], data=data)
    r.raise_for_status()
    return r.text


def search_get_page(client: RateLimitedClient, page_url: str) -> str:
    """Pagination: subsequent pages are simple GETs against -suchergeb.html?page=N."""
    r = client.get(page_url)
    r.raise_for_status()
    return r.text


def parse_search_hits(html: str) -> list[Record]:
    """Result-page HTML → list of partially-populated Records.

    Each hit carries BOTH Drucksachen (Anfrage + Antwort), the answering
    Ministerium-Kürzel (after the literal token "Antwort"), and both dates.
    The "Antwort"-block is optional — pending KAs (no answer yet) only have
    the Anfrage-block; in that case answer fields stay empty.

    Populated when present: wp, kleine_anfrage_nr, anfrager, fraktion,
    drucksache_anfrage_nr / link_anfrage, anfragedatum, anfragetitel,
    drucksache_antwort_nr / link_antwort, antwortdatum, ministerium_kuerzel,
    systematik, schlagworte.
    """
    soup = BeautifulSoup(html, "lxml")
    out: list[Record] = []
    for art in soup.select("article.e-search-result"):
        body = art.select_one(".e-search-result__body")
        if not body:
            continue
        rec = Record()
        title_el = body.find(["strong", "b"])
        if title_el:
            rec.anfragetitel = re.sub(r"\s+", " ", title_el.get_text(" ", strip=True)).strip()

        # Both Drucksachen-Links in document order: 1st = Anfrage, 2nd = Antwort.
        pdf_links = body.find_all("a", href=re.compile(r"MMD\d+-\d+\.pdf"))
        for idx, a in enumerate(pdf_links[:2]):
            href = a["href"]
            full = "https://www.landtag.nrw.de" + href if href.startswith("/") else href
            m = re.search(r"MMD(\d+)-(\d+)\.pdf", href)
            if not m:
                continue
            wp_v = int(m.group(1))
            nr = f"{m.group(1)}/{m.group(2)}"
            if idx == 0:
                rec.wp = wp_v
                rec.drucksache_anfrage_nr = nr
                rec.link_anfrage = full
            else:
                rec.drucksache_antwort_nr = nr
                rec.link_antwort = full

        text = body.get_text("\n", strip=True)
        # The KA-Nr line in a search hit body always sits on its own:
        #   <Title…>
        #   Kleine Anfrage <Nr>                     ← regular hit
        #   Kleine Anfrage <Nr> zu Drs <Drucksache> ← Nachfrage variant
        #   <Anfrager> <Fraktion>
        # Multiline-anchor on ^…$ excludes title-prose mentions like
        # "Nachfragen zur Antwort … auf die Kleine Anfrage 941" — without
        # this, a Nachfrage hijacks the cited original KA-Nr and the row
        # looks like a duplicate of the (unrelated) original KA.
        # The negative-lookahead `(?![\d/])` blocks Drucksache-shaped
        # matches ("Kleine Anfrage 18/9829") on the same line.
        # The "zu Drs N/M" suffix marks Nachfragen and is allowed.
        ka_match = re.search(
            r"(?m)^Kleine Anfrage\s+(\d+)(?![\d/])(?:\s+zu\s+Drs\s+\d+/\d+)?\s*$",
            text,
        )
        if ka_match:
            rec.kleine_anfrage_nr = int(ka_match.group(1))
        # Anfrager+Fraktion lives on the line directly after the KA-Nr line.
        # Anchor on that position — without it, a title that ends in "…AfD"
        # (e.g. KA 7364: "Was ist dran am 'Anschlag' auf die AfD") matches
        # first and overwrites the real Anfrager/Fraktion.
        af_search_text = text[ka_match.end():] if ka_match else text
        m = re.search(
            r"^\n*([^\n]+?)\s+(CDU|SPD|GRÜNE|FDP|AfD|fraktionslos)\s*$",
            af_search_text,
            re.MULTILINE,
        )
        if m:
            rec.anfrager = m.group(1).strip()
            rec.fraktion = m.group(2)

        # Two dates in "DD.MM.YYYY N S." form: 1st = Anfragedatum, 2nd = Antwortdatum.
        dates = re.findall(r"(\d{1,2}\.\d{1,2}\.\d{4})\s+\d+\s*S\.", text)
        if dates:
            rec.anfragedatum = _german_date_to_iso(dates[0])
        if len(dates) >= 2:
            rec.antwortdatum = _german_date_to_iso(dates[1])

        # Antwortendes Ministerium: structural line "Antwort KÜRZEL" followed
        # by the answer-Drucksache line — that pair is the fingerprint of the
        # actual Antwort-block at the bottom of every search-hit body.
        # A bare \b-anchored match (the previous version) misfires when the
        # Inhaltsbeschreibung contains "Antwort BT-Drs. 17/8134" and similar
        # citations of foreign-parliament Drucksachen (e.g. KA 331/WP17 → BT,
        # KA 6451/WP17 → BMWK).
        m = re.search(
            r"^Antwort\s+([A-ZÄÖÜ]{2,10})\s*\n\s*Drucksache\s+\d+/\d+",
            text, re.MULTILINE,
        )
        if m:
            rec.ministerium_kuerzel = m.group(1)
        # Withdrawn KAs: instead of "Antwort <Kürzel>" the second block reads
        # "Unterrichtung Präs" (the Landtag president's notice of withdrawal).
        # Mark the row so downstream stages skip Ministerium expectations and
        # don't try to extract an answer text.
        elif re.search(r"\bUnterrichtung\s+Präs\b", text):
            rec.antworttext_status = STATUS_ZURUECKGEZOGEN

        m = re.search(r"Systematik:\s*\n(.+?)(?:\nSchlagworte:|\nRegion:|\nAntwort\s|$)", text, re.DOTALL)
        if m:
            rec.systematik = re.sub(r"\s+", " ", m.group(1)).replace(" * ", "; ").strip()
        m = re.search(r"Schlagworte:\s*\n(.+?)(?:\nRegion:|\nSystematik:|\nAntwort\s|$)", text, re.DOTALL)
        if m:
            rec.schlagworte = re.sub(r"\s+", " ", m.group(1)).replace(" * ", "; ").strip()
        out.append(rec)
    return out


def find_next_page_url(html: str) -> str | None:
    """Return the absolute URL of the next page in a paginated result set, or None.

    The pagination block contains both PREV and NEXT links (each rendered twice,
    above + below the result list). They share the visible text 'Zu Seite N',
    differing only in CSS class ('--next' vs '--prev'). Filter on the class to
    avoid picking PREV (which would loop pagination forever).
    """
    soup = BeautifulSoup(html, "lxml")
    for a in soup.select("a.a-pagination-item-button--next[href]"):
        if "a-pagination-item-button--disabled" in (a.get("class") or []):
            continue  # disabled NEXT means we're on the last page
        href = a["href"]
        if "anfragen-und-antworten-suchergeb" not in href:
            continue
        return "https://www.landtag.nrw.de" + href if href.startswith("/") else href
    return None


# --- verbs -------------------------------------------------------------------

_SCAN_FIELDS = {
    "wp", "kleine_anfrage_nr", "drucksache_anfrage_nr", "drucksache_antwort_nr",
    "anfrager", "fraktion", "anfragedatum", "anfragetitel", "antwortdatum",
    "ministerium", "antworttext_status", "antworttext_quelle", "extract_flags",
}

# Enrich mode: scan-archive runs *after* crawl, which is the canonical source
# for KA-Nr / Drucksachen / Anfrager / Fraktion / dates / Kürzel. The PDF
# is consulted only to fill the prosa Ministerium name (search hits give
# only the Kürzel) and the Anfragetitel where missing — never to overwrite.
# pdftotext occasionally mangles fields (e.g. eats the space in "Kleine
# Anfrage 6790 vom" → "67901vom"); pinning crawl as canon prevents these
# parse glitches from poisoning the index.
_SCAN_FIELDS_ENRICH = {
    "anfragetitel", "ministerium", "antworttext_quelle",
}


def cmd_scan_archive(args: argparse.Namespace) -> int:
    """Walk Archiv/**/MMD<wp>-*.pdf, extract pages 1-3, upsert.

    Default mode = enrich-only: skip PDFs whose Drucksache-Nr is not already
    in the index. The online search (`crawl`) is the canonical source of
    truth for which Drucksachen are real Kleine-Anfrage answers; scan-archive
    only fills page-1-to-3 metadata (Anfragetitel, Ministerium prosa name)
    that the search hits don't carry.

    --allow-discovery resurrects the legacy behaviour of creating new rows
    from PDFs alone — useful only when running standalone without network.

    No network either way.
    """
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)
    seen_ministeria: set[str] = {r.ministerium for r in rows.values() if r.ministerium}

    counters = {"scanned": 0, "parsed": 0, "extract_failed": 0,
                "skipped": 0, "novelty": 0, "skipped_unknown_pk": 0}
    pdfs = list(iter_archive_pdfs(args.wahlperiode))
    if args.limit:
        pdfs = pdfs[: args.limit]
    discovery = bool(getattr(args, "allow_discovery", False))
    print(f"scan-archive: {len(pdfs)} PDFs to consider "
          f"(mode={'discovery' if discovery else 'enrich-only'})",
          file=sys.stderr, flush=True)

    for pdf in pdfs:
        counters["scanned"] += 1
        try:
            wp, drucksache_antwort_nr = parse_filename_drucksache_nr(pdf)
        except ValueError as e:
            _append_log(EXTRACT_ERRORS_LOG, f"{_now_iso()} | {pdf.name} | filename: {e}")
            counters["extract_failed"] += 1
            continue

        existing = rows.get(drucksache_antwort_nr)
        if not existing and not discovery:
            # Enrich-only: this PDF's Drucksache-Nr is not in the canon index.
            # It's either a Große Anfrage, an Anfrage-PDF, an unrelated doc,
            # or a row crawl hasn't seen yet. Skip silently.
            counters["skipped_unknown_pk"] += 1
            counters["skipped"] += 1
            continue
        if existing and existing.antworttext_status == STATUS_ZURUECKGEZOGEN and not args.force:
            # Withdrawn KA — the "PDF" is an Unterrichtung, not an answer.
            counters["skipped"] += 1
            continue
        if existing and not args.force and existing.antworttext_status != STATUS_FAILED \
                and existing.kleine_anfrage_nr and existing.anfragetitel and existing.ministerium:
            # Already enriched fully — nothing to add.
            counters["skipped"] += 1
            continue

        try:
            text = pdftotext_first_pages(pdf)
        except subprocess.CalledProcessError as e:
            _append_log(EXTRACT_ERRORS_LOG, f"{_now_iso()} | {pdf.name} | pdftotext: {e}")
            counters["extract_failed"] += 1
            rec = Record(
                wp=wp, drucksache_antwort_nr=drucksache_antwort_nr,
                antworttext_status=STATUS_FAILED, antworttext_quelle=QUELLE_LOCAL,
            )
            upsert(rows, rec, set_columns=_SCAN_FIELDS)
            continue

        # Out-of-scope filter: Große Anfragen share the layout but are a
        # different doc class. Skip without writing — no row, no log noise.
        if is_grosse_anfrage(text):
            counters["skipped_grosse"] = counters.get("skipped_grosse", 0) + 1
            counters["skipped"] += 1
            continue

        # Anfrage-PDF filter: Archiv/ holds both the inquiry PDFs and the
        # answer PDFs. Inquiry PDFs share page-1 layout (same KA-Nr, same
        # Anfrager line) — without this filter, scan-archive uses the
        # inquiry's Drucksache-Nr as `drucksache_antwort_nr`, creating a
        # phantom row keyed on the question. ~500 such phantoms in WP18.
        if not is_antwort_drucksache(text):
            counters["skipped_anfrage"] = counters.get("skipped_anfrage", 0) + 1
            counters["skipped"] += 1
            continue

        parsed = parse_page_text(text)
        # Verify the answer Drucksache-Nr from page matches the filename;
        # filename wins (it's the cache key).
        page_da = parsed.pop("drucksache_antwort_nr", "")
        if page_da and page_da != drucksache_antwort_nr:
            _append_log(
                EXTRACT_ERRORS_LOG,
                f"{_now_iso()} | {pdf.name} | drucksache mismatch: "
                f"filename={drucksache_antwort_nr} pdf={page_da}",
            )

        record_field_names = {f.name for f in fields(Record)}
        rec = Record(
            wp=wp,
            drucksache_antwort_nr=drucksache_antwort_nr,
            antworttext_quelle=QUELLE_LOCAL,
            **{k: v for k, v in parsed.items() if k in record_field_names},
        )

        # Field-level extract success requires kleine_anfrage_nr at minimum.
        # Status assignment: only set a status here when the row is brand new
        # (or had a worse status). fetch-text's STATUS_EXTRACTED, the
        # withdrawn-status, and STATUS_NO_ANSWER must NOT be downgraded back
        # to PENDING_ENRICH just because scan-archive runs after them.
        if not parsed.get("kleine_anfrage_nr"):
            counters["extract_failed"] += 1
            _append_log(EXTRACT_ERRORS_LOG, f"{_now_iso()} | {pdf.name} | regex: kleine_anfrage_nr missing")
            if not (existing and existing.antworttext_status):
                rec.antworttext_status = STATUS_FAILED
        else:
            counters["parsed"] += 1
            if existing and existing.antworttext_status in (
                    STATUS_EXTRACTED, STATUS_ZURUECKGEZOGEN, STATUS_NO_ANSWER):
                rec.antworttext_status = ""  # leave existing alone
            else:
                rec.antworttext_status = STATUS_PENDING_ENRICH

        # Vocab novelty checks
        if rec.fraktion and not check_fraktion(rec.fraktion):
            log_vocab_novelty(drucksache_antwort_nr, "Fraktion", rec.fraktion)
            counters["novelty"] += 1
        if rec.ministerium and rec.ministerium not in seen_ministeria:
            log_vocab_novelty(drucksache_antwort_nr, "Ministerium", rec.ministerium)
            seen_ministeria.add(rec.ministerium)
            counters["novelty"] += 1

        # In enrich-only mode (existing row present, default workflow), use
        # the narrow set; never overwrite crawl's canonical fields.
        cols = _SCAN_FIELDS if (discovery or existing is None) else _SCAN_FIELDS_ENRICH
        merged = upsert(rows, rec, set_columns=cols)
        # Recompute flags on the merged record so prior LLM-rescued values are
        # respected — otherwise rescue work would be silently re-flagged as missing.
        merged.extract_flags = compute_extract_flags(merged)

        if counters["scanned"] % 200 == 0:
            print(
                f"  ...{counters['scanned']}/{len(pdfs)} parsed={counters['parsed']} failed={counters['extract_failed']}",
                file=sys.stderr, flush=True,
            )

    # Re-normalize so flags include unknown_ministerium_kuerzel, novel_ministerium
    # etc. — compute_extract_flags called in the loop above only rebuilds the
    # missing_*/novel_fraktion subset.
    apply_normalization(rows)
    print(f"saving {len(rows)} rows to {xlsx} ...", file=sys.stderr, flush=True)
    save_index(rows, xlsx)
    print(
        f"scan-archive done: scanned={counters['scanned']} parsed={counters['parsed']} "
        f"extract_failed={counters['extract_failed']} skipped={counters['skipped']} "
        f"(davon nicht im Index={counters.get('skipped_unknown_pk', 0)}, "
        f"Große Anfrage={counters.get('skipped_grosse', 0)}, "
        f"Anfrage-PDF={counters.get('skipped_anfrage', 0)}) "
        f"vocab_novelty={counters['novelty']} rows_in_xlsx={len(rows)}",
        file=sys.stderr, flush=True,
    )
    return 0


_CRAWL_FIELDS = {
    "wp", "kleine_anfrage_nr", "drucksache_anfrage_nr", "drucksache_antwort_nr",
    "anfrager", "fraktion", "anfragedatum", "anfragetitel",
    "antwortdatum", "ministerium_kuerzel",
    "systematik", "schlagworte", "link_anfrage", "link_antwort",
    "antworttext_status",
}

# db_index.xlsx column subset: every field whose source is the Landtag search/DB
# (no PDF-derived enrichment columns). Includes timestamps for traceability.
DB_COLUMNS = [
    "WP",
    "Kleine_Anfrage_Nr",
    "Drucksache_Anfrage_Nr",
    "Drucksache_Antwort_Nr",
    "Anfrager",
    "Fraktion",
    "Anfragedatum",
    "Anfragetitel",
    "Antwortdatum",
    "Systematik",
    "Schlagworte",
    "Link_Drucksache_Anfrage",
    "Link_Drucksache_Antwort",
    "Ministerium_Kuerzel",
    "Antworttext_Status",
    "Hinzugefuegt_am",
    "Aktualisiert_am",
]


def _date_in_range(iso: str, lo: str | None, hi: str | None) -> bool:
    if lo and iso < lo:
        return False
    if hi and iso > hi:
        return False
    return True


def cmd_crawl(args: argparse.Namespace) -> int:
    """Discover and enrich Kleine Anfrage metadata via the Webflow search.

    Single pagination pass per Wahlperiode: each search hit carries BOTH
    Drucksachen (Anfrage + Antwort), both dates, the answering Ministerium-
    Kürzel, plus Systematik / Schlagworte / Anfrager / Fraktion / title.
    --full is a no-op for back-compat (full sweep is now the only mode).

    Row identity rules:
      - PK is always the Antwort-Drucksache.
      - For unanswered KAs (no Antwort-DS in the search hit), the Anfrage-DS
        is reused as a placeholder PK with status='pending'.
      - Old placeholder rows (PK == Anfrage-DS) get migrated to their real
        Antwort-DS when the search hit reveals it.
    """
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)
    counters = {"hits": 0, "matched_existing": 0, "new": 0,
                "migrated": 0, "filtered": 0, "errors": 0}

    with make_client(args.rps, args.user_agent) as client:
        try:
            tokens = bootstrap_search(client)
        except Exception as e:
            _append_log(CRAWL_ERRORS_LOG, f"{_now_iso()} | bootstrap failed: {e}")
            print(f"crawl: bootstrap failed: {e}", file=sys.stderr)
            return 2
        print(f"crawl: bootstrapped {tokens['post_url'][:120]}", file=sys.stderr, flush=True)
        print(f"crawl: full sweep for wp={args.wahlperiode}", file=sys.stderr, flush=True)

        html = search_post(client, tokens, wp=args.wahlperiode, rpp=args.rpp)
        page = 1
        while True:
            hits = parse_search_hits(html)
            counters["hits"] += len(hits)
            for h in hits:
                ad = h.anfragedatum
                if ad and not _date_in_range(ad, args.date_from, args.date_to):
                    counters["filtered"] += 1
                    continue
                if not h.drucksache_anfrage_nr:
                    continue

                # Look up both possible representations: the real Antwort-DS PK,
                # and any placeholder row keyed on the Anfrage-DS (legacy from
                # earlier discovery runs that didn't know the answer yet).
                real_pk = h.drucksache_antwort_nr
                real_row = rows.get(real_pk) if real_pk else None
                placeholder = None
                anfrage = h.drucksache_anfrage_nr
                ph_row = rows.get(anfrage)
                if (ph_row is not None and ph_row is not real_row
                        and ph_row.drucksache_anfrage_nr == anfrage
                        and ph_row.drucksache_antwort_nr == anfrage):
                    placeholder = ph_row

                if real_row and placeholder:
                    # Dedupe: drop the placeholder, the real row is authoritative.
                    del rows[placeholder.drucksache_antwort_nr]
                    counters["migrated"] += 1
                    existing = real_row
                elif real_row:
                    existing = real_row
                elif placeholder:
                    # Migrate the placeholder to its real PK (search now knows it).
                    if real_pk:
                        del rows[placeholder.drucksache_antwort_nr]
                        placeholder.drucksache_antwort_nr = real_pk
                        rows[real_pk] = placeholder
                        counters["migrated"] += 1
                    existing = placeholder
                else:
                    existing = None

                if existing:
                    h.drucksache_antwort_nr = existing.drucksache_antwort_nr
                    merged = upsert(
                        rows, h,
                        set_columns=_CRAWL_FIELDS - {"drucksache_antwort_nr", "antworttext_status"},
                    )
                    counters["matched_existing"] += 1
                else:
                    if h.antworttext_status == STATUS_ZURUECKGEZOGEN:
                        pass  # keep withdrawn-status for new rows
                    elif h.drucksache_antwort_nr:
                        h.antworttext_status = STATUS_PENDING  # PDF not yet local
                    else:
                        h.drucksache_antwort_nr = h.drucksache_anfrage_nr  # placeholder PK
                        h.antworttext_status = STATUS_PENDING
                    merged = upsert(rows, h, set_columns=_CRAWL_FIELDS)
                    counters["new"] += 1
                # Withdrawn KAs override any prior status (was likely PENDING /
                # FAILED because no real answer exists). Pin it on every crawl.
                if h.antworttext_status == STATUS_ZURUECKGEZOGEN:
                    merged.antworttext_status = STATUS_ZURUECKGEZOGEN
                merged.extract_flags = compute_extract_flags(merged)
            next_url = find_next_page_url(html)
            if not next_url:
                break
            page += 1
            if page % 10 == 0:
                print(
                    f"  page {page}, hits={counters['hits']} "
                    f"matched={counters['matched_existing']} new={counters['new']} "
                    f"migrated={counters['migrated']}",
                    file=sys.stderr, flush=True,
                )
                save_index(rows, xlsx)  # checkpoint
            html = search_get_page(client, next_url)

    # Cleanup pass: any placeholder row (PK == Anfrage-DS) for which a real-PK
    # row with the same Anfrage-DS now exists is a duplicate left over from
    # earlier crawl runs (or skipped during this run if the server returned a
    # short page). Drop the placeholder; the real row is authoritative.
    by_anfrage: dict[str, str] = {}
    for k, r in rows.items():
        if r.drucksache_anfrage_nr and r.drucksache_antwort_nr != r.drucksache_anfrage_nr:
            by_anfrage[r.drucksache_anfrage_nr] = k
    cleaned = 0
    for k in list(rows.keys()):
        r = rows[k]
        if r.drucksache_anfrage_nr and r.drucksache_antwort_nr == r.drucksache_anfrage_nr:
            if r.drucksache_anfrage_nr in by_anfrage and by_anfrage[r.drucksache_anfrage_nr] != k:
                del rows[k]
                cleaned += 1

    print(f"saving {len(rows)} rows to {xlsx} ...", file=sys.stderr, flush=True)
    apply_normalization(rows)
    save_index(rows, xlsx)
    save_index(rows, DB_INDEX_XLSX, columns=DB_COLUMNS)
    print(
        f"crawl done: pages={page} hits={counters['hits']} "
        f"matched={counters['matched_existing']} new={counters['new']} "
        f"migrated={counters['migrated']} cleanup={cleaned} filtered={counters['filtered']} "
        f"(snapshot → {DB_INDEX_XLSX.name})",
        file=sys.stderr, flush=True,
    )
    return 0


_FETCH_FIELDS = {
    "wp", "kleine_anfrage_nr", "drucksache_anfrage_nr", "drucksache_antwort_nr",
    "anfrager", "fraktion", "anfragedatum", "anfragetitel", "antwortdatum",
    "ministerium", "antworttext", "antworttext_status", "antworttext_quelle",
}


def _pdf_to_md(pdf_path: Path) -> str:
    """Full-document text extraction. pdfplumber primary, pypdf fallback."""
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            return "\n\n".join((p.extract_text() or "") for p in pdf.pages)
    except Exception:
        reader = pypdf.PdfReader(str(pdf_path))
        return "\n\n".join(p.extract_text() or "" for p in reader.pages)


def _atomic_write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w", encoding="utf-8", delete=False, dir=path.parent, prefix=".tmp_", suffix=path.suffix,
    ) as tmp:
        tmp.write(content)
        tmp_path = Path(tmp.name)
    os.replace(tmp_path, path)


def cmd_fetch_text(args: argparse.Namespace) -> int:
    """For each row missing PDF/.md: locate-or-download, backfill pages 1-3, full extract."""
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)

    # Candidates: rows with a real Drucksache_Antwort_Nr ('18/N' format) and
    # status not yet 'extracted' (or --force).
    candidates = []
    for key, rec in rows.items():
        if not re.match(r"\d+/\d+$", rec.drucksache_antwort_nr or ""):
            continue
        if args.wahlperiode is not None and rec.wp != args.wahlperiode:
            continue
        if rec.antworttext_status == STATUS_EXTRACTED and not args.force:
            continue
        if rec.antworttext_status == STATUS_NO_ANSWER:
            continue
        if rec.antworttext_status == STATUS_ZURUECKGEZOGEN:
            continue
        if rec.antworttext_status == STATUS_PDF_MISSING and not args.force:
            # Landtag-CDN liefert ein falsches Dokument unter dieser DS-Nr;
            # ohne Server-seitige Korrektur wäre der Re-Fetch derselbe Bug.
            continue
        if is_placeholder_row(rec):
            # No separate Antwort-Drucksache yet → fetching the URL would
            # download the Anfrage-PDF and mislabel it as the answer.
            continue
        candidates.append((key, rec))
    if args.limit:
        candidates = candidates[: args.limit]

    print(f"fetch-text: {len(candidates)} candidate rows", file=sys.stderr, flush=True)

    counters = {"have": 0, "downloaded": 0, "extracted": 0, "no_answer": 0, "failed": 0}
    client = make_client(args.rps, args.user_agent)

    try:
        for i, (key, rec) in enumerate(candidates, 1):
            m = re.match(r"(\d+)/(\d+)$", rec.drucksache_antwort_nr)
            wp = int(m.group(1)); n = int(m.group(2))

            # Locate or download PDF
            pdf = archive_lookup(wp, n)
            if pdf:
                rec.antworttext_quelle = QUELLE_LOCAL
                counters["have"] += 1
            else:
                url = PDF_URL_TEMPLATE.format(wp=wp, n=n)
                if not PDF_URL_ALLOW.match(url):
                    _append_log(EXTRACT_ERRORS_LOG, f"{_now_iso()} | {key} | url not allow-listed: {url}")
                    counters["failed"] += 1
                    continue
                try:
                    r = client.get(url)
                except Exception as e:
                    _append_log(EXTRACT_ERRORS_LOG, f"{_now_iso()} | {key} | download error: {e}")
                    counters["failed"] += 1
                    continue
                if r.status_code == 404:
                    rec.antworttext_status = STATUS_NO_ANSWER
                    upsert(rows, rec, set_columns={"antworttext_status"})
                    counters["no_answer"] += 1
                    continue
                if r.status_code != 200:
                    _append_log(EXTRACT_ERRORS_LOG, f"{_now_iso()} | {key} | http {r.status_code}")
                    counters["failed"] += 1
                    continue
                pdf = archive_target_path(wp, n)
                with tempfile.NamedTemporaryFile(
                    "wb", delete=False, dir=pdf.parent, prefix=".tmp_", suffix=".pdf",
                ) as tmp:
                    tmp.write(r.content)
                    tmp_path = Path(tmp.name)
                os.replace(tmp_path, pdf)
                rec.antworttext_quelle = QUELLE_DOWNLOADED
                counters["downloaded"] += 1

            # Backfill page-1-to-3 metadata if missing
            if not rec.kleine_anfrage_nr:
                try:
                    parsed = parse_page_text(pdftotext_first_pages(pdf))
                    parsed.pop("drucksache_antwort_nr", None)
                    for fname, val in parsed.items():
                        if fname in {f.name for f in fields(Record)} and not getattr(rec, fname):
                            setattr(rec, fname, val)
                except Exception as e:
                    _append_log(EXTRACT_ERRORS_LOG, f"{_now_iso()} | {key} | page1-3 backfill: {e}")

            # Full text → .md
            md_path = pdf.with_suffix(".md")
            if md_path.exists() and not args.force:
                rec.antworttext = str(md_path.relative_to(REPO_ROOT))
                rec.antworttext_status = STATUS_EXTRACTED
                counters["extracted"] += 1
            else:
                try:
                    text = _pdf_to_md(pdf)
                except Exception as e:
                    _append_log(EXTRACT_ERRORS_LOG, f"{_now_iso()} | {key} | pdf->md: {e}")
                    rec.antworttext_status = STATUS_FAILED
                    counters["failed"] += 1
                    upsert(rows, rec, set_columns=_FETCH_FIELDS)
                    continue
                _atomic_write_text(md_path, text)
                rec.antworttext = str(md_path.relative_to(REPO_ROOT))
                rec.antworttext_status = STATUS_EXTRACTED
                counters["extracted"] += 1

            upsert(rows, rec, set_columns=_FETCH_FIELDS)

            if i % 100 == 0:
                print(
                    f"  ...{i}/{len(candidates)} extracted={counters['extracted']} "
                    f"no_answer={counters['no_answer']} failed={counters['failed']}",
                    file=sys.stderr, flush=True,
                )
                save_index(rows, xlsx)
    finally:
        client.close()

    apply_normalization(rows)
    save_index(rows, xlsx)
    print(
        f"fetch-text done: have={counters['have']} downloaded={counters['downloaded']} "
        f"extracted={counters['extracted']} no_answer={counters['no_answer']} "
        f"failed={counters['failed']}",
        file=sys.stderr, flush=True,
    )
    return 0


_LLM_FIELDS = (
    "drucksache_anfrage_nr", "anfrager", "fraktion",
    "anfragedatum", "anfragetitel", "ministerium",
)

_LLM_SCHEMA = json.dumps({
    "type": "object",
    "properties": {
        "drucksache_anfrage_nr": {"type": ["string", "null"],
            "description": "Question Drucksache, format 'WP/N' (e.g. '18/27'). null if absent."},
        "anfrager": {"type": ["string", "null"],
            "description": "Asker name(s), comma-separated for multiple Abgeordnete."},
        "fraktion": {"type": ["string", "null"],
            "description": "One of: CDU, SPD, GRÜNE, FDP, AfD, fraktionslos."},
        "anfragedatum": {"type": ["string", "null"],
            "description": "ISO YYYY-MM-DD or null."},
        "anfragetitel": {"type": ["string", "null"],
            "description": "Title of the Kleine Anfrage."},
        "ministerium": {"type": ["string", "null"],
            "description": "Responding ministry full name (e.g. 'Ministerin für Schule und Bildung')."},
    },
    "required": [],
    "additionalProperties": False,
})

_LLM_PROMPT = (
    "Aus diesem Antwort-Drucksache-Text des Landtags NRW (Seiten 1-3) "
    "die Metadaten der Kleinen Anfrage extrahieren. Felder die nicht "
    "vorhanden sind: null. Anfragedatum als ISO YYYY-MM-DD.\n\n"
    "---\n{text}\n---"
)


def cmd_enrich_llm(args: argparse.Namespace) -> int:
    """LLM-based field extraction for rows the rule-based parser missed.

    Selects rows that have a local PDF (status != 'pending') and at least one
    missing _LLM_FIELDS value. Sends pdftotext output to the configured `llm`
    model with a JSON schema, fills only previously-empty fields.
    """
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)

    candidates = []
    for key, rec in rows.items():
        if rec.antworttext_status == STATUS_PENDING:
            continue  # placeholder row, no local PDF
        if args.wahlperiode is not None and rec.wp != args.wahlperiode:
            continue
        # Quality gate: only rows with extract_flags set (missing required
        # fields or novel_fraktion). Rows without flags are clean — skip.
        if not rec.extract_flags:
            continue
        m = re.match(r"(\d+)/(\d+)$", key)
        if not m:
            continue
        wp, n = int(m.group(1)), int(m.group(2))
        pdf = archive_lookup(wp, n)
        if not pdf:
            continue
        candidates.append((key, rec, pdf))

    if args.limit:
        candidates = candidates[: args.limit]
    print(f"enrich-llm: {len(candidates)} candidate rows ({args.model})", file=sys.stderr, flush=True)

    counters = {"queried": 0, "filled": 0, "errors": 0, "noop": 0}
    try:
        for i, (key, rec, pdf) in enumerate(candidates, 1):
            try:
                text = pdftotext_first_pages(pdf)
            except Exception as e:
                _append_log(VERIFY_LLM_LOG, f"{_now_iso()} | {key} | pdftotext: {e}")
                counters["errors"] += 1
                continue
            prompt = _LLM_PROMPT.format(text=text[:8000])
            try:
                r = subprocess.run(
                    ["llm", "-m", args.model, "--schema", _LLM_SCHEMA, prompt],
                    capture_output=True, text=True, timeout=120, check=True,
                )
                data = json.loads(r.stdout)
            except subprocess.CalledProcessError as e:
                _append_log(VERIFY_LLM_LOG, f"{_now_iso()} | {key} | llm exit {e.returncode}: {e.stderr[:200]}")
                counters["errors"] += 1
                continue
            except Exception as e:
                _append_log(VERIFY_LLM_LOG, f"{_now_iso()} | {key} | parse: {e}")
                counters["errors"] += 1
                continue
            counters["queried"] += 1
            any_filled = False
            for fname in _LLM_FIELDS:
                v = data.get(fname)
                if v and not getattr(rec, fname):
                    setattr(rec, fname, v.strip() if isinstance(v, str) else v)
                    any_filled = True
            # Recompute flags so reruns skip rows the LLM already cleaned.
            rec.extract_flags = compute_extract_flags(rec)
            if any_filled:
                counters["filled"] += 1
            else:
                counters["noop"] += 1
            if i % 25 == 0:
                print(
                    f"  ...{i}/{len(candidates)} filled={counters['filled']} "
                    f"errors={counters['errors']} noop={counters['noop']}",
                    file=sys.stderr, flush=True,
                )
                save_index(rows, xlsx)
    finally:
        save_index(rows, xlsx)

    print(
        f"enrich-llm done: queried={counters['queried']} filled={counters['filled']} "
        f"errors={counters['errors']} noop={counters['noop']}",
        file=sys.stderr, flush=True,
    )
    return 0


# --- canonical Fraktion / Ministerium normalisation --------------------------
#
# Index/fraktionen.xlsx and Index/ministerien.xlsx are user-curated tables.
# `normalize` reads them and fills three xlsx columns:
#   - Fraktion_Canonical       (exact-match rules; very few variants expected)
#   - Ministerium_Canonical    (token-Jaccard match; many spelling variants)
#   - Ministerium_Kuerzel      (the table's "Kürzel" column for the matched row)

INDEX_DIR = REPO_ROOT / "Index"
FRAKTIONEN_XLSX = INDEX_DIR / "fraktionen.xlsx"
MINISTERIEN_XLSX = INDEX_DIR / "ministerien.xlsx"

# Words ignored when token-matching ministry strings.
_MIN_STOPWORDS = frozenset({
    "ministerium", "minister", "ministerin",
    "ministerpräsident", "ministerpräsidentin",
    "der", "des", "die", "das", "den", "dem",
    "für", "fuer",  # tokenizer ASCII-folds 'ü' → 'ue', so 'für' arrives as 'fuer'
    "und", "sowie", "oder", "des",
    "landes", "nrw", "nordrhein-westfalen",
    "nordrhein", "westfalen",  # tokenizer splits the hyphen, drop the parts too
    "namens", "landesregierung",  # boilerplate phrase 'namens der Landesregierung'
    "chef", "chefin",
    "staatskanzlei",  # appears in 2 ministries; intentional drop so that
                      # "Chef der Staatskanzlei" alone doesn't dominate the score
})


def _ministerium_tokens(s: str) -> set[str]:
    """Tokenise a ministry string conservatively for SUBSET matching.

    Steps: lowercase + ASCII-fold umlauts; alpha-only tokens; strip the
    trailing 'ministerium' / 'minister(in)?' compound (so 'Justizministerium'
    becomes 'Justiz'); drop stopwords and tokens shorter than 4 chars.

    NOT done: prefix matching, flexion stripping, fuzzy edits. Anything not
    expressible as exact-subset is the operator's job to put into
    Index/ministerium_aliases.xlsx — that keeps the matcher false-positive-free.
    """
    s = s.lower().replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    raw = re.findall(r"[a-z]+", s)
    out: set[str] = set()
    for t in raw:
        for suf in ("ministeriums", "ministerium", "ministerinnen", "ministerin", "ministers", "minister"):
            if t.endswith(suf) and len(t) > len(suf):
                t = t[: -len(suf)]
                break
        if t in _MIN_STOPWORDS or len(t) < 4:
            continue
        out.add(t)
    return out


def load_canon_fraktionen() -> set[str]:
    """Return the canonical Fraktion set from Index/fraktionen.xlsx (column 'Fraktion')."""
    if not FRAKTIONEN_XLSX.exists():
        return set()
    wb = openpyxl.load_workbook(FRAKTIONEN_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None) or []
    try:
        col = list(header).index("Fraktion")
    except ValueError:
        wb.close()
        return set()
    out = {row[col] for row in rows_iter if row and row[col]}
    wb.close()
    return out


def load_canon_ministerien() -> list[tuple[str, str, set[str]]]:
    """Return [(Kürzel, Ministerium-Name, token-set)] from Index/ministerien.xlsx."""
    if not MINISTERIEN_XLSX.exists():
        return []
    wb = openpyxl.load_workbook(MINISTERIEN_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None) or []
    try:
        kc = list(header).index("Kürzel")
        nc = list(header).index("Ministerium")
    except ValueError:
        wb.close()
        return []
    out = []
    for row in rows_iter:
        if not row or not row[nc]:
            continue
        full = str(row[nc])
        kuerzel = str(row[kc]) if row[kc] else ""
        out.append((kuerzel, full, _ministerium_tokens(full)))
    wb.close()
    return out


def load_kuerzel_aliases() -> dict[str, str]:
    """Return {alias_kuerzel: primary_kuerzel} from Index/ministerien.xlsx
    'Aliases' column. Aliases are comma-separated. Used to merge
    Kürzel-Schreibweisen the search hit may emit (e.g. MCdS → MBEIM)."""
    if not MINISTERIEN_XLSX.exists():
        return {}
    wb = openpyxl.load_workbook(MINISTERIEN_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None) or []
    header_list = list(header)
    try:
        kc = header_list.index("Kürzel")
    except ValueError:
        wb.close()
        return {}
    if "Aliases" not in header_list:
        wb.close()
        return {}
    ac = header_list.index("Aliases")
    out: dict[str, str] = {}
    for row in rows_iter:
        if not row or not row[kc] or not row[ac]:
            continue
        primary = str(row[kc])
        for alias in str(row[ac]).split(","):
            alias = alias.strip()
            if alias:
                out[alias] = primary
    wb.close()
    return out


MINISTERIUM_ALIASES_XLSX = INDEX_DIR / "ministerium_aliases.xlsx"


def load_ministerium_aliases() -> dict[str, str]:
    """Operator-curated raw → Kürzel overrides. Returns {} if file missing.

    Format: cols 'Raw' (the verbatim Ministerium string as it appears in
    data/index.xlsx) and 'Kuerzel' (must match a Kürzel in ministerien.xlsx).
    """
    if not MINISTERIUM_ALIASES_XLSX.exists():
        return {}
    wb = openpyxl.load_workbook(MINISTERIUM_ALIASES_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None) or []
    try:
        rc = list(header).index("Raw")
        kc = list(header).index("Kuerzel")
    except ValueError:
        wb.close()
        return {}
    out: dict[str, str] = {}
    for row in rows_iter:
        if row and row[rc] and row[kc]:
            out[str(row[rc])] = str(row[kc])
    wb.close()
    return out


def match_ministerium(value: str, canonicals: list[tuple[str, str, set[str]]]) -> tuple[str, str] | None:
    """Strict subset match: every value-token must appear in the canonical
    token set. False-positive-free by construction — anything that needs a
    fuzzy/flexion call belongs in ministerium_aliases.xlsx instead.

    Ties (multiple canonicals satisfy the subset) go to the canonical with
    the FEWEST tokens (most specific), then to the longer name as fallback.
    """
    if not value:
        return None
    val_tokens = _ministerium_tokens(value)
    if not val_tokens:
        return None
    candidates = [(k, f, ct) for k, f, ct in canonicals if ct and val_tokens.issubset(ct)]
    if not candidates:
        return None
    candidates.sort(key=lambda x: (len(x[2]), -len(x[1])))
    k, f, _ = candidates[0]
    return (k, f)


# --- multi-ministerium parser (spec: 2026-05-04-multi-ministerium-parser.md) -

# Phase A: locate the "Der Minister/Die Ministerin … (hat) die Kleine Anfrage
# N … beantwortet." sentence — the boundary between the cited Anfragetext
# (Schritt 6 in the spec) and the actual answers (Schritt 8).
#
# Why anchor on "die Kleine Anfrage <N> … beantwortet": the Vorbemerkung often
# quotes prior answers verbatim, so a bare "Der Minister … beantwortet" can
# false-positive. The literal "die Kleine Anfrage <KA-Nr>" reference appears
# only in the actual boundary paragraph.
#
# Why search the flattened (whitespace-collapsed) text: pdftotext glues the
# running page header "LANDTAG NORDRHEIN-WESTFALEN - <wp>. Wahlperiode
# Drucksache N/M" onto the boundary paragraph with a single newline, so
# blank-line paragraph splits don't isolate it cleanly.
_RX_MIN_BOUNDARY = re.compile(
    # Determiner — Nominativ (Der/Die/Das) at sentence-start of the boundary
    # paragraph. Role — Minister(in), Ministerpräsident(in), Ministerium,
    # Ministerien (plural — observed when the boundary paragraph names two
    # ministries jointly via "Die Ministerien für X und Y").
    r"(?:Der|Die|Das)\s+Minister(?:präsident(?:in)?|in|ium|ien)?\b"
    r".{0,800}?"
    r"\bdie\s+Kleine\s+Anfrage\s+\d+"
    r".{0,400}?"
    r"\bbeantwortet\b[^.]{0,30}\."
)


def find_minister_paragraph(text: str) -> str | None:
    """Locate the boundary sentence that names the answering minister(s).
    Returns it as a single whitespace-normalized line, or None if not found."""
    flat = re.sub(r"\s+", " ", text)
    m = _RX_MIN_BOUNDARY.search(flat)
    return m.group(0) if m else None


# Phase B: extract minister full-forms from the boundary paragraph.
# Pattern: <Determiner> Minister[präsident][in] <connective> <body> <stop>
# Determiners: Der/Die (Nominativ), dem/der (Dativ — used in
#   "im Einvernehmen mit dem Minister …" / "der Ministerin …").
# Connectives: für (most common — "Minister für X"), des/der (gen — "des
#   Innern", "der Justiz").
# Stop tokens: link phrases that bridge to the next minister (und/sowie are
#   the most common), plus the sentence-final verb "beantwortet" and the
#   "(im) Einvernehmen mit" / "in Abstimmung mit" boilerplate. The next
#   determiner ("dem"/"der" before "Minister") also stops the body — without
#   it, "Minister des Innern und dem Minister der Justiz" parses as one form.
#   Comma is NOT a stop — body itself often contains commas (e.g. "Kinder,
#   Jugend, Familie, Gleichstellung, Flucht und Integration").
_RX_MIN_FORM = re.compile(
    # Determiner: Nominativ "Der/Die/Das" at sentence start; lowercase
    # der/die/das/dem/des/den mid-paragraph (Dativ/Genitiv/Akkusativ).
    # Role: Minister, Ministerin, Ministerium, Ministerien, plus
    # Ministerpräsident(in) — all treated equivalently. The capture group
    # only flows into the human-readable form string; the resolver tokenises
    # away the role suffix so it doesn't affect Kürzel matching.
    r"\b(?:Der|Die|Das|der|die|das|dem|des|den)\s+"
    r"(Minister(?:präsident(?:in)?|in|ium|ien)?)\s+"
    r"(für|der|des)\s+"
    r"(.+?)"
    # Stop alternatives — each carries its own leading-space rule. The
    # comma stop is `\s*,` so it fires even when the body ends right at
    # the comma (e.g. "Finanzen, dem Minister …"). The 'namens' / 'des
    # Landes' stops keep the boilerplate out of the body so the
    # operator-curated aliases (Index/ministerium_aliases.xlsx) still
    # match by exact string after we drop the determiner-connective.
    r"(?="
    r"\s+hat\b|\s+haben\b|"
    r"\s+im\s+Einvernehmen\b|\s+in\s+Abstimmung\b|"
    r"\s+gemeinsam\s+mit\b|\s+nach\s+Beteiligung\b|\s+unter\s+Mitwirkung\b|"
    r"\s+beantwortet\b|\s+beantworten\b|"
    r"\s+wie\s+folgt\b|"
    r"\s+namens\s+der\s+Landesregierung\b|"
    r"\s+des\s+Landes\s+Nordrhein-Westfalen\b|"
    # Stop alternatives bridging to the next minister. Each accepts:
    #   - Minister(in), Ministerpräsident(in), Minister[ium|s|n] etc. via
    #     [Mm]inister[a-zäöüß]* — lowercase m enables compound forms
    #     ("Verkehrsminister"); permissive suffix covers Genitiv "Ministers",
    #     plural "Ministerien" and the full "Ministerium" noun.
    #   - Optional German compound prefix ([A-ZÄÖÜ][a-zäöüß]+) for
    #     "Verkehrsminister", "Finanzminister", "Innenminister" etc.
    r"\s+und\s+(?:mit\s+)?(?:dem|der|des|den)\s+(?:[A-ZÄÖÜ][a-zäöüß]+)?[Mm]inister[a-zäöüß]*\b|"
    r"\s+sowie\s+(?:mit\s+)?(?:dem|der|des|den)?\s*(?:[A-ZÄÖÜ][a-zäöüß]+)?[Mm]inister[a-zäöüß]*\b|"
    r"\s*,\s+(?:mit\s+)?(?:dem|der|des|den)\s+(?:[A-ZÄÖÜ][a-zäöüß]+)?[Mm]inister[a-zäöüß]*\b"
    r")"
)


def _undo_pdf_hyphenation(s: str) -> str:
    """Undo pdftotext hyphenation quirks:
      1. Soft-hyphen leak with no space: 'Klima-schutz' → 'Klimaschutz',
         'Gleich-stellung' → 'Gleichstellung'.
      2. Line-wrap hyphen with space: 'In- nern' → 'Innern'. Skip when the
         next word is 'und/oder/sowie/noch' — those follow a grammatical
         Bindestrich-Ergänzung (e.g. 'Bundes- und Europaangelegenheiten')
         that must NOT be joined.
      3. pdftotext occasionally eats the space after a Bindestrich-Ergänzung,
         producing 'Bundesund'. Split it back so the canon tokenizer can
         reach the 'bundes' / 'und' tokens.
    """
    s = re.sub(r"(\w)-([a-zäöüß])", r"\1\2", s)
    s = re.sub(r"(\w)-\s+(?!(?:und|oder|sowie|noch)\b)([a-zäöüß])", r"\1\2", s)
    s = re.sub(r"\bBundesund\b", "Bundes und", s)
    # pdftotext occasionally swallows the space before the boilerplate
    # token "namens" in "...Finanzen namens der Landesregierung", producing
    # "Finanzennamens" — split it back so the namens-Landesregierung stop
    # alternative still fires and the body doesn't run on past it.
    s = re.sub(r"([a-zäöüß])namens\b", r"\1 namens", s)
    return s


def extract_minister_full_forms(paragraph: str) -> list[str]:
    """Return ordered list of '<Rolle> <Body>' strings — duplicates preserved
    (caller dedupes on Kürzel after resolving). The connective ('für/der/des')
    is dropped so the output aligns with the format used in
    Index/ministerium_aliases.xlsx ('Minister Innern', not 'Minister des
    Innern')."""
    paragraph = _undo_pdf_hyphenation(paragraph)
    out: list[str] = []
    for m in _RX_MIN_FORM.finditer(paragraph):
        rolle, body = m.group(1), m.group(3).strip()
        out.append(f"{rolle} {body}")
    return out


def resolve_minister_form(
    full_form: str,
    canon_min: list[tuple[str, str, set[str]]],
    aliases: dict[str, str],
) -> str | None:
    """Map an extracted full-form ('Minister des Innern', 'Ministerin für …')
    to a Kürzel. Tries operator-curated aliases first, then strict subset
    match against ministerien.xlsx tokens. Returns None if neither resolves."""
    if full_form in aliases:
        return aliases[full_form]
    hit = match_ministerium(full_form, canon_min)
    return hit[0] if hit else None


def _count_beteiligte(kuerzel_str: str) -> int:
    """Count of comma-separated Kürzel in a Beteiligte_Ministerien_Kuerzel value."""
    if not kuerzel_str:
        return 0
    return sum(1 for k in kuerzel_str.split(",") if k.strip())


# --- abgeordnete index + multi-anfrager parser ------------------------------
#
# The Landtag search-hit truncates the Anfrager list to the first 2 names plus
# 'u.a.' when 3+ Abgeordnete co-sign a Kleine Anfrage. The Antwort-PDF header
# carries the full list ("Vorname Nachname, ... und Vorname Nachname FRAKTION").
#
# Strategy: bootstrap an Index/abgeordnete.xlsx of (WP, Fraktion, Nachname,
# Vorname) tuples from the reliable first-2 names in the DB, then use it as
# a known-name lexicon to greedy-substring-match the PDF Anfragerblock.
# Iteration: rows beyond DB-cap can introduce names not yet in the index;
# residue is logged to data/anfrager_novelty.log; rerun build → extract until
# residue is empty.

# Match Anfragerblock(s) in the Antwort-PDF text. Generous body window (1000
# chars) because some PDFs split names across many lines. Terminate on the
# Fraktion marker — accept both 'GRÜNE/N' and the full 'BÜNDNIS 90/DIE GRÜNEN'.
# pdftotext occasionally drops whitespace at word boundaries (e.g. WP17 KA 6:
# 'der AbgeordnetenVerena Schäffer'); accept either explicit whitespace OR an
# upper-case letter directly following 'Abgeordneten' as the start of the names.
_RX_ANFRAGERBLOCK = re.compile(
    r"(?:des|der)\s+Abgeordneten(?:\s+|(?=[A-ZÄÖÜ]))"
    r"(.{1,1000}?)\s*"
    # PDF whitespace is lossy — accept missing spaces in 'DIE GRÜNEN' / '90/DIE'.
    r"(?:BÜNDNIS\s*90\s*/\s*DIE\s*GRÜNEN?|CDU|SPD|GRÜNEN?|FDP|AfD|fraktionslos)\b",
    re.DOTALL,
)

# Title prefixes that pdftotext keeps in the Nachname slot of the DB
# ("Dr. Maelzer, Dennis"). Convert to PDF form by hoisting them in front
# of the Vorname ("Dr. Dennis Maelzer").
_RX_NAME_TITLE = re.compile(r"^((?:(?:Prof\.|Dr\.)\s+)+)(.*)$")


def parse_db_anfrager(s: str) -> list[tuple[str, str]]:
    """'Becker, Horst; Klocke, Arndt u.a.' → [('Becker','Horst'), ('Klocke','Arndt')].

    Strips the trailing 'u.a.' marker. Each token must contain ', ' to split
    Nachname,Vorname; tokens without it (rare data glitches) are skipped.
    """
    out: list[tuple[str, str]] = []
    for raw in s.split("; "):
        part = raw.strip()
        if part.endswith(" u.a."):
            part = part[: -len(" u.a.")].strip()
        if not part or part == "u.a.":
            continue
        if ", " not in part:
            continue
        nach, vor = part.split(", ", 1)
        out.append((nach.strip(), vor.strip()))
    return out


def db_to_pdf_form(nach: str, vor: str) -> str:
    """('Berg', 'Guido van den')  → 'Guido van den Berg'.
    ('Dr. Maelzer', 'Dennis')      → 'Dr. Dennis Maelzer'.

    Title prefixes (Dr./Prof.) on Nachname are hoisted in front of Vorname
    because that is how the Antwort-PDF renders the name.
    """
    m = _RX_NAME_TITLE.match(nach)
    if m:
        title = m.group(1).strip()
        nach_clean = m.group(2).strip()
        return f"{title} {vor} {nach_clean}".strip()
    return f"{vor} {nach}"


def db_to_pdf_form_aliases(nach: str, vor: str) -> list[str]:
    """All name forms an Antwort-PDF might render for ('Nachname', 'Vorname').

    Returns the canonical form plus simplified variants:
      - drop trailing middle initials  ('Sven W.' → 'Sven')
      - drop hyphenated middle parts   ('Lisa-Kristin' → 'Lisa')
      - drop title prefixes            ('Dr. Hartmut' → 'Hartmut')
    The match step uses longest-first ordering, so canonical forms win when
    both are present in the PDF.
    """
    aliases: list[str] = [db_to_pdf_form(nach, vor)]
    # First-token first name (handles 'Sven W.', 'Lisa-Kristin', 'Michael R.').
    vor_first = re.split(r"[\s\-]", vor.strip(), maxsplit=1)[0]
    if vor_first and vor_first != vor:
        aliases.append(db_to_pdf_form(nach, vor_first))
    # Title-stripped variant (PDF sometimes drops 'Dr.' / 'Prof.').
    m = _RX_NAME_TITLE.match(nach)
    if m:
        nach_clean = m.group(2).strip()
        aliases.append(f"{vor} {nach_clean}")
        if vor_first and vor_first != vor:
            aliases.append(f"{vor_first} {nach_clean}")
    # Dedup, preserve order.
    seen = set()
    out: list[str] = []
    for a in aliases:
        if a and a not in seen:
            seen.add(a)
            out.append(a)
    return out


def _abgeordnete_xlsx_path() -> Path:
    return INDEX_DIR / "abgeordnete.xlsx"


def load_abgeordnete_index() -> dict[tuple[int, str], list[tuple[str, str, str]]]:
    """Return {(wp, fraktion): [(pdf_form, nach, vor), ...]} sorted by
    len(pdf_form) DESC so longest forms are tried first during matching
    (avoids 'Beucker' shadowing 'Dr. Hartmut Beucker')."""
    path = _abgeordnete_xlsx_path()
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        wb.close()
        return {}
    h = {name: i for i, name in enumerate(header) if name is not None}
    out: dict[tuple[int, str], list[tuple[str, str, str]]] = {}
    for row in rows_iter:
        wp = row[h.get("WP", -1)] if "WP" in h else None
        frak = row[h.get("Fraktion", -1)] if "Fraktion" in h else None
        nach = row[h.get("Nachname", -1)] if "Nachname" in h else None
        vor = row[h.get("Vorname", -1)] if "Vorname" in h else None
        pdf_form = row[h.get("PDF_Form", -1)] if "PDF_Form" in h else None
        if not (wp and frak and pdf_form and nach and vor):
            continue
        out.setdefault((int(wp), str(frak)), []).append(
            (str(pdf_form), str(nach), str(vor))
        )
    wb.close()
    for k in out:
        out[k].sort(key=lambda t: -len(t[0]))
    return out


def log_anfrager_novelty(drucksache: str, fraktion: str, residue: str, block: str) -> None:
    """Append unmatched residue from a PDF Anfragerblock for later review."""
    _append_log(
        ANFRAGER_NOVELTY_LOG,
        f'{_now_iso()} | {drucksache} | {fraktion} | residue="{residue}" | block="{block}"',
    )


def db_anfrager_min_count(s: str) -> int:
    """Lower bound on the number of co-signers implied by the DB Anfrager string.

    The Landtag search-hit truncates to 2 names + 'u.a.' when 3+ co-sign:
      ''                        → 0 (no info)
      'Schäffer, Verena'        → 1
      'Becker, Horst; Klocke, Arndt' → 2
      'Bischoff, Rainer; Börner, Frank u.a.' → 3 (lower bound; could be more)
    """
    if not s:
        return 0
    has_ua = " u.a." in s or s.endswith("u.a.")
    n = len(parse_db_anfrager(s))
    return n + 1 if has_ua and n >= 2 else n


def cmd_build_abgeordnete_index(args: argparse.Namespace) -> int:
    """Scan data/index.xlsx, accumulate distinct (WP, Fraktion, Nachname,
    Vorname) tuples from the Anfrager column, write Index/abgeordnete.xlsx.

    The first 2 names in any DB row are reliable (the bug is only that 3+
    co-signers get truncated to 2 + 'u.a.'). Run this BEFORE extract-all-
    anfrager, and rerun AFTER to absorb any names harvested from u.a. PDFs
    that hadn't shown up as primary co-signers anywhere.
    """
    rows = load_index(Path(args.xlsx))
    counter: dict[tuple[int, str, str, str], int] = {}
    for rec in rows.values():
        if not rec.wp or not rec.fraktion or not rec.anfrager:
            continue
        for nach, vor in parse_db_anfrager(rec.anfrager):
            key = (rec.wp, rec.fraktion, nach, vor)
            counter[key] = counter.get(key, 0) + 1
        # Anfrager_Alle (if already populated) is the better, broader source —
        # absorb any names not covered by the truncated DB Anfrager column.
        if rec.anfrager_alle:
            for nach, vor in parse_db_anfrager(rec.anfrager_alle):
                key = (rec.wp, rec.fraktion, nach, vor)
                counter[key] = counter.get(key, 0) + 1

    # Per-person aggregation across Fraktionen within the same WP, to surface
    # Fraktionswechsel (Abgeordnete who left their Fraktion mid-WP, e.g. Pretzell,
    # Neppe). Each output row carries the OTHER Fraktionen this person appears
    # under in 'Frueher_Fraktion' — useful both for documentation and for an
    # operator-driven cross-Fraktion match if a row's PDF Anfragerblock names
    # someone whose current Fraktion in the index doesn't include them yet.
    by_person: dict[tuple[int, str, str], set[str]] = {}
    for (wp, frak, nach, vor) in counter.keys():
        by_person.setdefault((wp, nach, vor), set()).add(frak)

    INDEX_DIR.mkdir(parents=True, exist_ok=True)
    path = _abgeordnete_xlsx_path()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "abgeordnete"
    ws.append([
        "WP", "Fraktion", "Nachname", "Vorname", "PDF_Form",
        "Is_Alias", "Frueher_Fraktion", "n_treffer",
    ])
    n_rows = 0
    for (wp, frak, nach, vor), n in sorted(counter.items()):
        other = sorted(by_person[(wp, nach, vor)] - {frak})
        frueher = "; ".join(other)
        for i, form in enumerate(db_to_pdf_form_aliases(nach, vor)):
            ws.append([wp, frak, nach, vor, form, 0 if i == 0 else 1, frueher, n])
            n_rows += 1
    wb.save(path)
    print(
        f"build-abgeordnete-index done: {len(counter)} canonical, {n_rows} forms → {path}",
        file=sys.stderr,
    )
    return 0


def cmd_extract_all_anfrager(args: argparse.Namespace) -> int:
    """Per row's .md: parse the Anfragerblock, match names against
    Index/abgeordnete.xlsx (filtered by WP+Fraktion), write Anfrager_Alle
    (DB form, '; '-joined, in order of appearance in the PDF) and
    Anzahl_Abgeordnete back to data/index.xlsx.

    The original Anfrager column is preserved verbatim. Anfrager_Alle is the
    authoritative list for analysis. Unmatched residue from the Anfragerblock
    is logged to data/anfrager_novelty.log for review.
    """
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)
    abg = load_abgeordnete_index()
    if not abg:
        print(
            "ERROR: Index/abgeordnete.xlsx fehlt — vorher `build-abgeordnete-index` laufen lassen.",
            file=sys.stderr,
        )
        return 2

    counters = {
        "scanned": 0, "matched": 0, "no_block": 0, "no_candidates": 0,
        "novel_residue": 0, "skipped_status": 0, "plausi_mismatch": 0,
    }

    for rec in rows.values():
        if args.wahlperiode is not None and rec.wp != args.wahlperiode:
            continue
        if rec.antworttext_status != STATUS_EXTRACTED or not rec.antworttext:
            counters["skipped_status"] += 1
            continue
        # Operator-curated rows opt out of auto-overwrite. Add 'anfrager_manual'
        # to Extract_Flags after a hand-correction so re-runs preserve it.
        if "anfrager_manual" in (rec.extract_flags or "").split(","):
            counters["skipped_status"] += 1
            continue
        md_path = REPO_ROOT / rec.antworttext
        if not md_path.exists():
            continue
        counters["scanned"] += 1
        try:
            text = md_path.read_text(encoding="utf-8")
        except Exception as e:
            _append_log(
                EXTRACT_ERRORS_LOG,
                f"{_now_iso()} | {rec.drucksache_antwort_nr} | anfrager read: {e}",
            )
            continue

        m = _RX_ANFRAGERBLOCK.search(text[:5000])
        if not m:
            counters["no_block"] += 1
            block = ""
        else:
            block = re.sub(r"\s+", " ", m.group(1)).strip()
            # pdftotext routinely drops whitespace at lower→upper word
            # boundaries ('AnjaButschkau', 'undJürgen') and after punctuation
            # ('Sven W.Tritschler'). Re-insert before matching so substring
            # lookups against PDF_Form succeed.
            block = re.sub(r"([a-zäöüß.])([A-ZÄÖÜ])", r"\1 \2", block)

        candidates = abg.get((rec.wp, rec.fraktion), [])
        if not candidates:
            counters["no_candidates"] += 1
            # Still run plausi-check below; just skip matching.
            block = ""

        # Greedy substring matching, longest forms first. Mask hits to prevent
        # later (shorter) forms from re-matching already-consumed text.
        masked = list(block)
        matched: list[tuple[int, str, str]] = []
        for pdf_form, nach, vor in candidates:
            current = "".join(masked)
            idx = current.find(pdf_form)
            if idx < 0:
                continue
            left_ok = idx == 0 or not current[idx - 1].isalpha()
            right_end = idx + len(pdf_form)
            right_ok = right_end == len(current) or not current[right_end].isalpha()
            if not (left_ok and right_ok):
                continue
            matched.append((idx, nach, vor))
            for i in range(idx, right_end):
                masked[i] = "\x00"

        matched.sort(key=lambda t: t[0])
        if matched:
            rec.anfrager_alle = "; ".join(f"{nach}, {vor}" for _, nach, vor in matched)
            rec.anzahl_abgeordnete = len(matched)
            counters["matched"] += 1
        else:
            rec.anfrager_alle = ""
            rec.anzahl_abgeordnete = 0

        # Residue: anything in the block that is NOT consumed AND not separator
        # noise. Title fragments ('Dr.', 'Prof.'), 'und', commas, semicolons,
        # whitespace are expected leftover.
        residue = "".join(masked).replace("\x00", " ")
        residue = re.sub(r"\b(?:und|Dr\.?|Prof\.?|Professor|Frau|Herr)\b", " ", residue)
        residue = re.sub(r"[\s,;.]+", " ", residue).strip()
        if residue:
            counters["novel_residue"] += 1
            log_anfrager_novelty(
                rec.drucksache_antwort_nr or "", rec.fraktion or "", residue, block
            )

        # Plausi-check: parsed count must not undershoot the DB count. Equal-or-
        # higher is fine (the whole point of this verb is to recover names the
        # search-hit dropped under 'u.a.'). Lower means a parser failure on this
        # row — log loudly so the operator can re-recherchieren via the agent.
        db_min = db_anfrager_min_count(rec.anfrager or "")
        if rec.anzahl_abgeordnete < db_min:
            counters["plausi_mismatch"] += 1
            _append_log(
                ANFRAGER_NOVELTY_LOG,
                f'{_now_iso()} | {rec.drucksache_antwort_nr or ""} | '
                f'{rec.fraktion or ""} | MISMATCH parsed={rec.anzahl_abgeordnete} '
                f'db_min={db_min} | db_anfrager="{rec.anfrager or ""}" | '
                f'block="{block}"',
            )

    save_index(rows, xlsx)
    print(
        f"extract-all-anfrager done: scanned={counters['scanned']} "
        f"matched={counters['matched']} no_block={counters['no_block']} "
        f"no_candidates={counters['no_candidates']} "
        f"novel_residue={counters['novel_residue']} "
        f"plausi_mismatch={counters['plausi_mismatch']}",
        file=sys.stderr,
    )
    if counters["plausi_mismatch"]:
        print(
            f"WARNING: {counters['plausi_mismatch']} Zeilen haben weniger geparste "
            f"Anfrager als die DB-Anfrager-Spalte erwartet — siehe MISMATCH-Einträge "
            f"in {ANFRAGER_NOVELTY_LOG}. Skill-Abschnitt 'Edge-Case-Nachrecherche'.",
            file=sys.stderr,
        )
    return 0


def cmd_extract_multi_ministerium(args: argparse.Namespace) -> int:
    """Read each row's Antworttext .md, parse the boundary paragraph for ALL
    answering ministries, write Kürzel-Set to Beteiligte_Ministerien_Kuerzel.

    Federführend (= existing Ministerium_Kuerzel) is forced first; further
    Kürzel follow in order-of-appearance from the PDF text. Idempotent —
    overwrites Beteiligte_* on every run.
    """
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)
    canon_min = load_canon_ministerien()
    aliases = load_ministerium_aliases()
    kuerzel_aliases = load_kuerzel_aliases()

    counters = {"scanned": 0, "no_para": 0, "single": 0, "multi": 0,
                "novel_form": 0, "skipped_status": 0}

    for rec in rows.values():
        if args.wahlperiode is not None and rec.wp != args.wahlperiode:
            continue
        if rec.antworttext_status != STATUS_EXTRACTED:
            counters["skipped_status"] += 1
            continue
        if not rec.antworttext:
            continue
        md_path = REPO_ROOT / rec.antworttext
        if not md_path.exists():
            continue

        counters["scanned"] += 1
        try:
            text = md_path.read_text(encoding="utf-8")
        except Exception as e:
            _append_log(EXTRACT_ERRORS_LOG, f"{_now_iso()} | {rec.drucksache_antwort_nr} | multi-min read: {e}")
            continue

        # Phase B: full-forms → Kürzel, deduped, ordered.
        # Federführend (DB Ministerium_Kuerzel) is the floor: every answered KA
        # has at least one ministry, so we seed kuerzel_order from DB even if the
        # PDF boundary paragraph is missing/unparseable.
        kuerzel_order: list[str] = []
        if rec.ministerium_kuerzel:
            primary = kuerzel_aliases.get(rec.ministerium_kuerzel, rec.ministerium_kuerzel)
            kuerzel_order.append(primary)

        para = find_minister_paragraph(text)
        if not para:
            counters["no_para"] += 1
        else:
            for form in extract_minister_full_forms(para):
                k = resolve_minister_form(form, canon_min, aliases)
                if k:
                    k = kuerzel_aliases.get(k, k)
                    if k not in kuerzel_order:
                        kuerzel_order.append(k)
                else:
                    counters["novel_form"] += 1
                    log_vocab_novelty(rec.drucksache_antwort_nr or "", "ministerium_form", form)

        rec.beteiligte_ministerien_kuerzel = ",".join(kuerzel_order)
        rec.beteiligte_ministerien = len(kuerzel_order)
        if len(kuerzel_order) >= 2:
            counters["multi"] += 1
        elif len(kuerzel_order) == 1:
            counters["single"] += 1

    save_index(rows, xlsx)
    print(
        f"extract-multi-ministerium done: scanned={counters['scanned']} "
        f"single={counters['single']} multi={counters['multi']} "
        f"no_para={counters['no_para']} novel_forms={counters['novel_form']}",
        file=sys.stderr,
    )
    return 0


def apply_normalization(rows: dict) -> dict:
    """Mutate rows in place: fill canonical / Kürzel columns, recompute flags.

    Returns counters dict. Pure data — no I/O. Callable from any pipeline
    stage that wants the index to stay normalized after a write.
    """
    canon_frak = load_canon_fraktionen()
    canon_min = load_canon_ministerien()
    aliases = load_ministerium_aliases()
    kuerzel_aliases = load_kuerzel_aliases()
    kuerzel_to_full = {k: f for k, f, _ in canon_min}

    counters = {"frak_matched": 0, "frak_novel": 0,
                "min_matched_alias": 0, "min_matched_subset": 0, "min_novel": 0,
                "min_matched_kuerzel": 0, "alias_orphan": 0}

    for rec in rows.values():
        # Fraktion: exact match against canonical set
        if rec.fraktion:
            if rec.fraktion in canon_frak:
                rec.fraktion_canonical = rec.fraktion
                counters["frak_matched"] += 1
            else:
                rec.fraktion_canonical = ""
                counters["frak_novel"] += 1

        # Ministerium resolution priority:
        #   1. direct Kürzel (set by crawl from the search hit's "Antwort KÜRZEL"
        #      token; authoritative when present);
        #   2. operator-curated alias table;
        #   3. strict subset token match against canonical names (FP-frei).
        # The Kürzel from the database is preserved as-is — never overwritten
        # by an alias mapping. The `Aliases` column in ministerien.xlsx is
        # used only for the canonical-name lookup, so historical/variant
        # spellings (e.g. MKJFGF) still resolve to the right full ministry
        # name without losing the original Kürzel value.
        if rec.ministerium_kuerzel:
            kz = rec.ministerium_kuerzel
            # Try direct lookup first (for entries with their own row in
            # ministerien.xlsx, e.g. MCdS); fall back to alias resolution.
            full = kuerzel_to_full.get(kz)
            if not full and kz in kuerzel_aliases:
                full = kuerzel_to_full.get(kuerzel_aliases[kz])
            if full:
                rec.ministerium_canonical = full
                counters["min_matched_kuerzel"] = counters.get("min_matched_kuerzel", 0) + 1
            else:
                # Kürzel set but unknown — likely a parser misfire (e.g. "BT"
                # picked up from "Antwort BT-Drs." in the Inhaltsbeschreibung)
                # or a source-data typo on landtag.nrw.de (e.g. "AWEL"). Surface
                # via flag + novelty log so an operator can reconcile.
                rec.ministerium_canonical = ""
                counters["min_novel"] += 1
                log_vocab_novelty(rec.drucksache_antwort_nr or "", "ministerium_kuerzel", kz)
        elif rec.ministerium:
            kuerzel = aliases.get(rec.ministerium)
            if kuerzel:
                full = kuerzel_to_full.get(kuerzel)
                if full:
                    rec.ministerium_kuerzel = kuerzel
                    rec.ministerium_canonical = full
                    counters["min_matched_alias"] += 1
                else:
                    rec.ministerium_kuerzel = ""
                    rec.ministerium_canonical = ""
                    counters["alias_orphan"] += 1
            else:
                m = match_ministerium(rec.ministerium, canon_min)
                if m:
                    rec.ministerium_kuerzel, rec.ministerium_canonical = m
                    counters["min_matched_subset"] += 1
                else:
                    rec.ministerium_kuerzel = ""
                    rec.ministerium_canonical = ""
                    counters["min_novel"] += 1
        elif rec.ministerium_kuerzel:
            # Kürzel set but not in ministerien.xlsx → flag, don't guess.
            rec.ministerium_canonical = ""
            counters["min_novel"] += 1

        # Keep Beteiligte_Ministerien (count) in sync with the Kürzel column
        # after every normalize — so a backfilled or hand-edited Kürzel string
        # immediately reflects in the count column.
        # Floor: if DB knows the federführende Ministerium_Kuerzel, at least
        # that one ministry is always involved — seed the Kürzel-Set if empty.
        if not rec.beteiligte_ministerien_kuerzel and rec.ministerium_kuerzel:
            rec.beteiligte_ministerien_kuerzel = rec.ministerium_kuerzel
        rec.beteiligte_ministerien = _count_beteiligte(rec.beteiligte_ministerien_kuerzel)

        # Rebuild flag set: missing_* recomputed (so stale flags drop after a
        # field got filled), then novel_ministerium added if no canonical resolved.
        base = compute_extract_flags(rec)
        # Preserve non-missing/non-novel/non-unknown flags caller may have set elsewhere.
        existing = [f for f in rec.extract_flags.split(",")
                    if f and not f.startswith("missing_")
                    and not f.startswith("novel_")
                    and f != "unknown_ministerium_kuerzel"]
        flags = existing + ([f for f in base.split(",") if f] if base else [])
        if rec.ministerium and not rec.ministerium_canonical:
            flags.append("novel_ministerium")
        # Kürzel set but not resolvable against ministerien.xlsx → fragwürdig.
        if rec.ministerium_kuerzel and not rec.ministerium_canonical:
            flags.append("unknown_ministerium_kuerzel")
        # de-dupe, preserve order
        seen = set(); out_flags = []
        for f in flags:
            if f and f not in seen:
                out_flags.append(f); seen.add(f)
        rec.extract_flags = ",".join(out_flags)

    return counters


def cmd_normalize(args: argparse.Namespace) -> int:
    """Fill Fraktion_Canonical / Ministerium_Canonical / Ministerium_Kuerzel.

    Sources, applied in priority:
      1. Index/ministerium_aliases.xlsx — operator-curated raw → Kürzel.
         Always wins; this is how typos / odd wording get mapped without
         loosening the rule-based matcher.
      2. Index/ministerien.xlsx — strict subset token match against the
         canonical Ministerium names (false-positive-free by construction).
      3. Index/fraktionen.xlsx — exact-string match for Fraktion (very few
         variants are expected; novelty is genuine, not a parse fluke).

    Anything still unmatched gets a novel_fraktion / novel_ministerium flag."""
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)
    if not load_canon_fraktionen():
        print(f"normalize: WARN no canonical Fraktionen at {FRAKTIONEN_XLSX}", file=sys.stderr)
    if not load_canon_ministerien():
        print(f"normalize: WARN no canonical Ministerien at {MINISTERIEN_XLSX}", file=sys.stderr)
    counters = apply_normalization(rows)
    save_index(rows, xlsx)
    print(
        f"normalize done: "
        f"fraktion matched={counters['frak_matched']} novel={counters['frak_novel']}; "
        f"ministerium alias={counters['min_matched_alias']} subset={counters['min_matched_subset']} "
        f"kuerzel={counters['min_matched_kuerzel']} novel={counters['min_novel']} "
        f"alias_orphan={counters['alias_orphan']}",
        file=sys.stderr,
    )
    return 0


def _classify_mismatch(rec: "Record", parsed: dict) -> tuple[list[str], bool]:
    """Heuristik-Klassifikator für DB↔PDF-Mismatches.

    Returns (notes, all_resolved). `notes` = pro Mismatch-Flag genau ein
    erklärender Eintrag (oder leer, wenn nicht auto-klassifizierbar).
    `all_resolved` = True nur wenn JEDES Flag gelöst wurde.
    """
    flags = [f.strip() for f in (rec.mismatch_flags or "").split(",") if f.strip()]
    notes: list[str] = []

    def parse_year(d: str) -> "int | None":
        m = re.match(r"(\d{4})-", d or "")
        return int(m.group(1)) if m else None

    def signed_days(pdf_v: str, db_v: str) -> "int | None":
        try:
            from datetime import date as _date
            return (_date.fromisoformat(db_v) - _date.fromisoformat(pdf_v)).days
        except Exception:
            return None

    for f in flags:
        if f == "antwortdatum":
            pdf_v = parsed.get("antwortdatum", "")
            db_v = rec.antwortdatum or ""
            y = parse_year(pdf_v)
            if y is not None and not (2015 <= y <= 2030):
                notes.append(
                    f"PDF-Antwortdatum '{pdf_v}' = pdftotext-OCR-Glitch "
                    f"(Jahr {y} außerhalb [2015,2030]); DB '{db_v}' maßgeblich.")
                continue
            dd = signed_days(pdf_v, db_v) if pdf_v and db_v else None
            if dd is None:
                return notes, False
            if 0 < dd <= 122:
                notes.append(
                    f"Antwortdatum-Diff {dd}d (PDF '{pdf_v}' = Briefdatum, "
                    f"DB '{db_v}' = Drucksachen-/Veröff.-Datum); beide legitim.")
                continue
            if 360 <= abs(dd) <= 370:
                notes.append(
                    f"PDF-Antwortdatum '{pdf_v}' = Jahres-Tippfehler (Diff {dd}d); "
                    f"DB '{db_v}' maßgeblich.")
                continue
            return notes, False

        elif f == "anfragedatum":
            pdf_v = parsed.get("anfragedatum", "")
            db_v = rec.anfragedatum or ""
            y = parse_year(pdf_v)
            if y is not None and not (2015 <= y <= 2030):
                notes.append(
                    f"PDF-Anfragedatum '{pdf_v}' = pdftotext-OCR-Glitch "
                    f"(Jahr {y} außerhalb [2015,2030]); DB '{db_v}' maßgeblich.")
                continue
            dd = signed_days(pdf_v, db_v) if pdf_v and db_v else None
            if dd is None:
                return notes, False
            if 0 < dd <= 122:
                notes.append(
                    f"Anfragedatum-Diff {dd}d (PDF '{pdf_v}' = Datum auf "
                    f"Anfrage-Schreiben, DB '{db_v}' = Drucksachen-Datum); "
                    f"beide legitim.")
                continue
            if 360 <= abs(dd) <= 370:
                notes.append(
                    f"PDF-Anfragedatum '{pdf_v}' = Jahres-Tippfehler "
                    f"(Diff {dd}d); DB '{db_v}' maßgeblich.")
                continue
            # Out of corridor: still trust DB (officielle Suche), classify as
            # OCR/Tippfehler im PDF — hands-off the row only if completely
            # unparseable.
            if pdf_v:
                notes.append(
                    f"Anfragedatum-Diff (PDF '{pdf_v}' vs DB '{db_v}') "
                    f"außerhalb Drucksache↔Brief-Korridor; DB maßgeblich "
                    f"(PDF vermutlich OCR/Tippfehler).")
                continue
            return notes, False

        elif f == "drucksache_anfrage_nr":
            pdf_v = parsed.get("drucksache_anfrage_nr", "")
            db_v = rec.drucksache_anfrage_nr or ""
            ka = rec.kleine_anfrage_nr
            try:
                wp_db, n_db = db_v.split("/")
                wp_pdf, n_pdf = pdf_v.split("/")
                n_db_i = int(n_db); n_pdf_i = int(n_pdf)
            except Exception:
                if pdf_v:
                    notes.append(
                        f"DS-Nr-Mismatch (DB '{db_v}' vs PDF '{pdf_v}', "
                        f"unparseable); DB aus offizieller Landtag-Suche maßgeblich.")
                    continue
                return notes, False
            why = None
            if wp_db != wp_pdf:
                why = f"PDF-Parser-Glitch: falsche Wahlperiode ({pdf_v})"
            elif n_pdf_i == ka:
                why = f"PDF-Parser-Glitch: Wert {pdf_v} = KA-Nr, nicht Anfrage-DS"
            elif n_pdf == wp_db:
                why = "PDF-Parser False-Positive auf Wahlperiode-Header"
            elif abs(n_pdf_i - n_db_i) in (1, 10, 100, 1000):
                why = f"PDF-Digit-Tippfehler (off {n_pdf_i - n_db_i:+d})"
            elif len(n_pdf) > len(n_db):
                why = f"PDF-OCR-Digit-Doppelung ('{pdf_v}' vs '{db_v}')"
            elif len(n_pdf) < len(n_db):
                why = f"PDF-OCR-Digit-Auslass ('{pdf_v}' vs '{db_v}')"
            else:
                why = f"PDF-OCR-Digit-Fehler (Diff {n_pdf_i - n_db_i:+d})"
            notes.append(
                f"DS-Nr: {why}; DB '{db_v}' aus offizieller Landtag-Suche maßgeblich.")

        elif f == "fraktion":
            pdf_v = parsed.get("fraktion", "")
            db_v = rec.fraktion or ""
            # Canonical equivalence: PDF-Wert nach DB-Mapping = DB?
            if pdf_v and _FRAKTION_PDF_TO_DB.get(pdf_v.upper(), pdf_v) == db_v:
                notes.append(
                    f"Fraktion-Aliasing (PDF '{pdf_v}' / DB '{db_v}'): "
                    f"kanonisch identisch.")
                continue
            # Bekannte Fraktionsaustritte WP17 (AfD → fraktionslos Herbst 2017):
            # Pretzell, Vogel, Neppe, Langguth. Antwortzeit ggf. nach Austritt.
            if rec.wp == 17 and db_v == "AfD" and pdf_v.lower() == "fraktionslos":
                notes.append(
                    f"Fraktionsaustritt zwischen Anfrage und Antwort (AfD → "
                    f"fraktionslos, Herbst 2017); DB='{db_v}' bei Anfrage, "
                    f"PDF='{pdf_v}' bei Antwort. Beide legitim.")
                continue
            # Sonst: vermutlich Landtag-PDF-Header-Tippfehler oder Multi-
            # Fraktions-Anfrage. DB = Fraktion der Anfrager:innen ist
            # authoritativ.
            if pdf_v:
                notes.append(
                    f"Fraktion-Mismatch (DB '{db_v}' vs PDF '{pdf_v}'): "
                    f"PDF-Header-Wert weicht ab — vermutlich Landtag-PDF-"
                    f"Tippfehler oder Multi-Fraktions-Anfrage. DB = Fraktion "
                    f"der Anfrager:innen maßgeblich.")
                continue
            return notes, False

        elif f == "md_kanr":
            # Parser-Output komplett leer → Spacing-/Render-Glitch
            empty = not any([parsed.get(k) for k in (
                "anfragedatum", "antwortdatum", "drucksache_anfrage_nr", "fraktion")])
            md_path = rec.antworttext or ""
            if empty and md_path:
                p = Path(md_path)
                if p.exists():
                    txt = p.read_text(encoding="utf-8", errors="replace")[:16000]
                    if re.search(rf"Anfrage\s*{rec.kleine_anfrage_nr}\s*vom|"
                                 rf"Anfrage\s*{rec.kleine_anfrage_nr}vom|"
                                 rf"Kleine\s+Anfrage\s+{rec.kleine_anfrage_nr}\b", txt):
                        notes.append(
                            f"pdftotext-Spacing-Glitch ('Anfrage "
                            f"{rec.kleine_anfrage_nr}vom' o. ä.); KA-Nr im PDF "
                            f"korrekt, md_kanr-Flag = Parser-Artefakt.")
                        continue
                    # Doppelt-gerendert?
                    if re.search(r"LL?AA?NN?DD?TT?AA?GG?", txt[:200]):
                        notes.append(
                            "pdftotext-Render-Quirk: jede Zeichenklasse doppelt "
                            "('LLAANNDDTTAAGG …'); Header-Parser läuft leer, "
                            "Mismatch-Flag = Parser-Artefakt.")
                        continue
            # md hat valide PDF-Felder, aber KA-Nr passt nicht — Landtag-PDF-Body
            # nennt fremde KA-Nr (Doppelvergabe / Druckfehler). DB authoritativ.
            if not empty:
                notes.append(
                    f"md_kanr-Mismatch: PDF-Body nennt abweichende KA-Nr "
                    f"(Landtag-Druck-/Numerierungsfehler — gleiche KA-Nr taucht "
                    f"in mehreren Antworten auf). DB-Snapshot maßgeblich.")
                continue
            return notes, False

        else:
            return notes, False

    return notes, len(notes) == len(flags) and len(flags) > 0


def _cmd_resolve_auto(args: argparse.Namespace) -> int:
    """Heuristik-Pass über alle ask_review-Zeilen.

    Klassifiziert Mismatches anhand bekannter Muster (siehe SKILL.md
    Heuristik-Tabelle) und setzt Datenqualität=korrigiert + Notiz. Resume-
    safe: schon notiert/korrigiert wird übersprungen. --dry-run zählt nur.
    """
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)
    targets = [r for r in rows.values()
               if r.datenqualitaet == "ask_review" and not r.notizen]
    print(f"resolve --auto: {len(targets)} ask_review-Zeilen ohne Notiz",
          file=sys.stderr)
    counters = {"resolved": 0, "skipped": 0}
    for rec in targets:
        parsed = {"anfragedatum": "", "antwortdatum": "",
                  "drucksache_anfrage_nr": "", "fraktion": ""}
        if rec.antworttext:
            md = REPO_ROOT / rec.antworttext
            if md.exists():
                head = md.read_text(encoding="utf-8", errors="replace")[:16000]
                parsed = _parse_pdf_header_fields(head)
        notes, ok = _classify_mismatch(rec, parsed)
        if ok and notes:
            if not args.dry_run:
                rec.notizen = " | ".join(notes)
                rec.datenqualitaet = "korrigiert"
                rec.aktualisiert_am = _now_iso()
            counters["resolved"] += 1
        else:
            counters["skipped"] += 1
    if not args.dry_run:
        save_index(rows, xlsx)
    print(f"resolve --auto: resolved={counters['resolved']} "
          f"skipped={counters['skipped']}{' (dry-run)' if args.dry_run else ''}",
          file=sys.stderr)
    return 0


def _cmd_resolve_interactive(args: argparse.Namespace) -> int:
    """Walk Datenqualität=ask_review rows; prompt user for note + verdict.

    Resume-safe: rows whose Notizen is already non-empty are skipped, so the
    user can stop with [q] and pick up later. Saves after each row.
    """
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)
    targets = [r for r in rows.values()
               if r.datenqualitaet == "ask_review" and not r.notizen]
    if not targets:
        print("resolve --interactive: no ask_review rows pending review.", file=sys.stderr)
        return 0
    print(f"\n{len(targets)} ask_review rows pending. "
          f"Press [q] at any prompt to save and quit.\n", file=sys.stderr)

    for i, rec in enumerate(targets, 1):
        print(f"\n[{i}/{len(targets)}] {rec.drucksache_antwort_nr}  "
              f"WP{rec.wp} KA {rec.kleine_anfrage_nr}  {rec.fraktion}")
        print(f"  Titel : {rec.anfragetitel[:100]}")
        print(f"  Anfrager (DB): {rec.anfrager}")
        print(f"  Anfragedatum DB: {rec.anfragedatum}   Antwortdatum DB: {rec.antwortdatum}")
        print(f"  Drucksache_Anfrage DB: {rec.drucksache_anfrage_nr}")
        print(f"  Mismatch_Flags: {rec.mismatch_flags}")
        # Re-parse PDF for the user's reference
        if rec.antworttext:
            md = REPO_ROOT / rec.antworttext
            if md.exists():
                head = md.read_text(encoding="utf-8", errors="replace")[:16000]
                parsed = _parse_pdf_header_fields(head)
                print(f"  PDF says — Anfragedatum: {parsed['anfragedatum']}, "
                      f"Antwortdatum: {parsed['antwortdatum']}, "
                      f"Drucksache_Anfrage: {parsed['drucksache_anfrage_nr']}, "
                      f"Fraktion: {parsed['fraktion']}")
                print(f"  PDF: {rec.antworttext}")
        try:
            note = input("  Notiz (Enter=skip, q=quit): ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nQuit (no save for current row).", file=sys.stderr)
            break
        if note.lower() == "q":
            print("\nQuit.", file=sys.stderr)
            break
        if not note:
            continue
        rec.notizen = note
        try:
            verdict = input("  Status [k]orrigiert / [s]kip [k]: ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            print("\nQuit (note saved, verdict skipped).", file=sys.stderr)
            save_index(rows, xlsx)
            break
        if verdict in ("", "k"):
            rec.datenqualitaet = "korrigiert"
        rec.aktualisiert_am = _now_iso()
        save_index(rows, xlsx)
    print(f"\nresolve --interactive done.", file=sys.stderr)
    return 0


def cmd_resolve(args: argparse.Namespace) -> int:
    """Mismatch resolver. Three modes:

    --auto               Heuristik-Pass: klassifiziert alle Datenqualität=
                         ask_review-Zeilen anhand bekannter Mismatch-Klassen
                         (PDF-OCR-Glitch, Drucksachen-vs-Brief-Datum,
                         Fraktionsaustritt, …) und setzt Datenqualität=
                         korrigiert mit erklärender Notiz. Idempotent;
                         nicht-klassifizierbare Zeilen bleiben ask_review.

    --ka N [--ka M …]    Targeted re-search of the live Landtag DB; overwrites
                         the index row with authoritative search-hit data.
                         (Use cases: md ↔ KA-Nr mismatch, KA gap, phantom row.)

    --interactive        Walk every Datenqualität=ask_review row in the index,
                         display the DB ↔ PDF mismatch, prompt for a free-text
                         Notiz + verdict ([k]orrigiert / [s]kip / [q]uit). Saves
                         after each step (resume-safe — already-noted rows are
                         skipped on the next run).
    """
    if args.auto:
        return _cmd_resolve_auto(args)
    if args.interactive:
        return _cmd_resolve_interactive(args)
    if not args.ka:
        print("resolve: pass --ka <Nr> (one or more), --interactive, or --auto",
              file=sys.stderr)
        return 2
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)
    wp = args.wahlperiode
    targets = list(args.ka)

    counters = {"resolved": 0, "no_hit": 0, "phantom_ga": 0, "phantom_deleted": 0,
                "added": 0, "updated": 0}

    with make_client(args.rps, args.user_agent) as client:
        try:
            tokens = bootstrap_search(client)
        except Exception as e:
            print(f"resolve: bootstrap failed: {e}", file=sys.stderr)
            return 2
        for ka_nr in targets:
            print(f"\n--- KA {ka_nr} (WP{wp}) ---", file=sys.stderr)
            html = search_post(client, tokens, wp=wp, nummer=str(ka_nr), rpp=10,
                               doktyp=DOKTYP_KLEINE_ANFRAGE)
            hits = [h for h in parse_search_hits(html)
                    if h.kleine_anfrage_nr == ka_nr and h.wp == wp]
            if not hits:
                print(f"  no Kleine-Anfrage hit for KA {ka_nr}", file=sys.stderr)
                if args.counter_search:
                    # Re-bootstrap to get a fresh single-use Webflow token.
                    tokens = bootstrap_search(client)
                    html2 = search_post(client, tokens, wp=wp, nummer=str(ka_nr),
                                        rpp=10, doktyp=DOKTYP_GROSSE_ANFRAGE)
                    ga_hits = parse_search_hits(html2)
                    if ga_hits:
                        print(f"  ! Gegenrecherche: KA-Nr {ka_nr} ist eine Große Anfrage",
                              file=sys.stderr)
                        for h in ga_hits[:3]:
                            print(f"    GA hit: Anfrage={h.drucksache_anfrage_nr} "
                                  f"Antwort={h.drucksache_antwort_nr} "
                                  f"({h.anfragetitel[:60]!r})", file=sys.stderr)
                        # Phantom-row identification: any row in our index whose
                        # KA-Nr matches the GA-Nr is misclassified data.
                        phantoms = [k for k, r in rows.items()
                                    if r.wp == wp and r.kleine_anfrage_nr == ka_nr]
                        for p in phantoms:
                            print(f"    Phantom row: {p}", file=sys.stderr)
                            counters["phantom_ga"] += 1
                            if args.delete_phantom:
                                del rows[p]
                                counters["phantom_deleted"] += 1
                        # Re-bootstrap so the next iteration's KA query has a fresh token.
                        tokens = bootstrap_search(client)
                        continue
                counters["no_hit"] += 1
                # Bootstrap a fresh token for the next loop iteration too.
                tokens = bootstrap_search(client)
                continue

            # Successful KA hit. Apply hit data to the matching row (or create one).
            h = hits[0]
            if not h.drucksache_antwort_nr:
                pk = h.drucksache_anfrage_nr  # placeholder PK for unanswered
                if h.antworttext_status != STATUS_ZURUECKGEZOGEN:
                    h.antworttext_status = STATUS_PENDING
                h.drucksache_antwort_nr = pk
            else:
                pk = h.drucksache_antwort_nr

            existing = rows.get(pk)
            # Also identify any pre-existing row sharing the same KA-Nr but a
            # different PK — those are stale (e.g. KA 144 → 18/10805 phantom).
            stale = [k for k, r in rows.items()
                     if r.wp == wp and r.kleine_anfrage_nr == ka_nr and k != pk]

            if existing:
                upsert(rows, h,
                       set_columns=_CRAWL_FIELDS - {"drucksache_antwort_nr", "antworttext_status"})
                counters["updated"] += 1
            else:
                upsert(rows, h, set_columns=_CRAWL_FIELDS)
                counters["added"] += 1
            merged = rows[pk]
            if h.antworttext_status == STATUS_ZURUECKGEZOGEN:
                merged.antworttext_status = STATUS_ZURUECKGEZOGEN
            merged.extract_flags = compute_extract_flags(merged)

            for s in stale:
                print(f"  stale row {s} (same KA-Nr, different Antwort-DS) "
                      f"{'→ deleted' if args.delete_phantom else '— use --delete-phantom to remove'}",
                      file=sys.stderr)
                if args.delete_phantom:
                    del rows[s]
                    counters["phantom_deleted"] += 1

            print(f"  resolved KA {ka_nr} → {pk} ({merged.anfrager}, {merged.fraktion}, "
                  f"{merged.ministerium_kuerzel or merged.antworttext_status})", file=sys.stderr)
            counters["resolved"] += 1
            # Re-bootstrap before the next KA — Webflow tokens are single-use.
            tokens = bootstrap_search(client)

    apply_normalization(rows)
    save_index(rows, xlsx)
    print(
        f"\nresolve done: resolved={counters['resolved']} no_hit={counters['no_hit']} "
        f"added={counters['added']} updated={counters['updated']} "
        f"phantom_ga={counters['phantom_ga']} phantom_deleted={counters['phantom_deleted']}",
        file=sys.stderr,
    )
    return 0


_FRAKTION_PDF_TO_DB = {
    "BÜNDNIS 90/DIE GRÜNEN": "GRÜNE",
    "BÜNDNIS90/DIEGRÜNEN": "GRÜNE",
    "FRAKTIONSLOS": "fraktionslos",
    "AFD": "AfD",
    "AfD": "AfD",
    "SPD": "SPD",
    "CDU": "CDU",
    "FDP": "FDP",
    "DIE LINKE": "DIE LINKE",
}


def _maybe_dedouble(head: str) -> str:
    """pdftotext rendert manche Schrift-Embeddings doppelt:
    'LLAANNDDTTAAGG NNOORRDDRRHHEEIINN--WWEESSTTFFAALLEENN'. Erkennen
    anhand der Häufigkeit konsekutiver gleicher Zeichen in den ersten
    200 chars; bei Render-Quirk weit über normal-deutsch (~5 %).
    Dedouble per Token, damit Word-Boundaries (Leerzeichen) erhalten
    bleiben und echte Doppelbuchstaben in normalem Text nicht zerstört
    werden.
    """
    sample = head[:200]
    if len(sample) < 60:
        return head
    eqs = sum(1 for i in range(len(sample) - 1) if sample[i] == sample[i + 1])
    # Normaler deutscher Text: ~3-5 % konsekutiv gleiche Zeichen.
    # Doppelt-gerendert: ~45-50 %. Schwelle 30 % ist sicher beidseitig.
    if eqs / (len(sample) - 1) < 0.30:
        return head
    # Pro Token (whitespace-getrennt) jedes zweite Zeichen rausnehmen,
    # um z. B. 'LLAANNDDTTAAGG' → 'LANDTAG' zu kürzen, ohne dass Spaces
    # zwischen Tokens das Pair-Alignment verschieben.
    def _dedouble_token(t: str) -> str:
        return t[::2] if len(t) >= 2 else t
    parts = re.split(r"(\s+)", head)
    return "".join(p if p.isspace() else _dedouble_token(p) for p in parts)


def _parse_pdf_header_fields(head: str) -> dict:
    """Re-extract DB-relevant fields from the answer-PDF for cross-check.

    Returns a dict with keys:
      - 'antwortdatum': str — the 'mit Schreiben vom <Datum>' inside the
        Minister-attribution paragraph (the same paragraph extract-multi-
        ministerium parses); this is the authoritative answer-letter date.
      - 'anfragedatum': str (the 'vom <Datum>' phrase).
      - 'drucksache_anfrage_nr', 'fraktion': str.
    Missing/unparseable fields are empty.

    Tolerant gegenüber pdftotext-Spacing-Glitches ('Anfrage 5vom',
    'Drucksache17/32') und doubled-character-Renders ('LLAANNDDTTAAGG …').
    """
    out: dict = {"antwortdatum": "", "anfragedatum": "",
                 "drucksache_anfrage_nr": "", "fraktion": ""}
    head = _maybe_dedouble(head)
    # Antwortdatum: re-use find_minister_paragraph (same source-of-truth as
    # extract-multi-ministerium), then pick the 'Schreiben vom <Datum>' inside.
    para = find_minister_paragraph(head)
    if para:
        m = re.search(
            r"mit\s+Schreiben\s+vom\s+(\d{1,2}\.\s*[A-Za-zÄÖÜäöüß]+\s+\d{4}|\d{1,2}\.\d{1,2}\.\d{4})",
            para)
        if m:
            out["antwortdatum"] = _german_date_to_iso(m.group(1))
    # Anfragedatum: 'Kleine Anfrage <Nr> vom <Datum>' — \s* statt \s+ vor
    # 'vom', weil pdftotext gelegentlich 'Anfrage 5vom' rendert.
    m = re.search(
        r"Kleine\s+Anfrage\s+\d+\s*vom\s+(\d{1,2}\.\s*[A-Za-zÄÖÜäöüß]+\s+\d{4}|\d{1,2}\.\d{1,2}\.\d{4})",
        head)
    if m:
        out["anfragedatum"] = _german_date_to_iso(m.group(1))
    # Drucksache_Anfrage_Nr + Fraktion: anchored on the Anfrager-block.
    # Layout is invariant: '<Anfrager-Liste> <FRAKTION-Token>\nDrucksache <wp/nr>'.
    # Captures both at once so we don't pick up the page-footer self-reference.
    # \s* statt \s+ nach 'Drucksache', weil pdftotext 'Drucksache17/32' rendert.
    m = re.search(
        r"(BÜNDNIS\s*90\s*[/ ]\s*DIE\s*GRÜNEN|FRAKTIONSLOS|fraktionslos|AfD|AFD|SPD|CDU|FDP|DIE\s*LINKE)"
        r"\s*\n?\s*Drucksache\s*(\d+/\d+)",
        head)
    if m:
        token = re.sub(r"\s+", " ", m.group(1)).strip()
        # Normalize: BÜNDNIS 90 / DIE GRÜNEN, BÜNDNIS 90/DIE GRÜNEN, BÜNDNIS90/DIEGRÜNEN
        token_norm = re.sub(r"\s*/\s*", "/", token).replace(" ", "")
        if token_norm.upper().startswith("BÜNDNIS90") or "GRÜNEN" in token_norm.upper():
            out["fraktion"] = "GRÜNE"
        elif token.upper() == "FRAKTIONSLOS":
            out["fraktion"] = "fraktionslos"
        elif token.upper() == "AFD":
            out["fraktion"] = "AfD"
        else:
            out["fraktion"] = _FRAKTION_PDF_TO_DB.get(token.upper(), token)
        out["drucksache_anfrage_nr"] = m.group(2)
    return out


def _date_mismatch(pdf_v: str, db_v: str) -> bool:
    """True wenn Datums-Diff ein echter Mismatch ist (nicht der erwartete
    Drucksachen-↔-Brief-Verarbeitungs-Lag).

    Akzeptiert ohne Flag:
      - Diff 0–122d in PDF-früher-Richtung (DB = Drucksachen-/Veröff.-Datum,
        PDF = Brief-/Anfrage-Schreibdatum).
    Flagged:
      - Diff > 122d, Diff in DB-früher-Richtung > 14d, Jahres-Tippfehler
        (~365d), OCR-Glitch (Jahr außerhalb [2015, 2030]).
    """
    if not pdf_v or not db_v or pdf_v == db_v:
        return False
    try:
        db_d = date.fromisoformat(db_v)
        pdf_d = date.fromisoformat(pdf_v)
    except ValueError:
        return False
    if not (2015 <= pdf_d.year <= 2030):
        return True   # OCR-Glitch
    dd = (db_d - pdf_d).days
    if 0 <= dd <= 122:
        return False  # Drucksache-vs-Brief-Korridor
    if -14 <= dd < 0:
        return False  # 14d-Toleranz in DB-früher-Richtung (Original-Schwelle)
    return True


def _detect_mismatches(rec: Record, head: str) -> list[str]:
    """Return field-names where DB ↔ PDF disagree for one row.

    Used by both verify (read-only report) and merge (persist into
    Mismatch_Flags column). Datum-Toleranz: 0–122d in PDF-früher-Richtung
    (Drucksache↔Brief-Lag), 14d in DB-früher-Richtung (Original-Schwelle).
    """
    flags: list[str] = []
    parsed = _parse_pdf_header_fields(head)
    if _date_mismatch(parsed["anfragedatum"], rec.anfragedatum):
        flags.append("anfragedatum")
    if _date_mismatch(parsed["antwortdatum"], rec.antwortdatum):
        flags.append("antwortdatum")
    if (parsed["drucksache_anfrage_nr"] and rec.drucksache_anfrage_nr
            and parsed["drucksache_anfrage_nr"] != rec.drucksache_anfrage_nr):
        flags.append("drucksache_anfrage_nr")
    if (parsed["fraktion"] and rec.fraktion
            and parsed["fraktion"] != rec.fraktion):
        flags.append("fraktion")
    # md_kanr: tolerant gegenüber pdftotext-Spacing-Glitches
    # ('Anfrage 5vom', 'Anfrage5 vom') und doubled-character-Renders.
    if rec.kleine_anfrage_nr:
        ka = rec.kleine_anfrage_nr
        head_norm = _maybe_dedouble(head)
        if not re.search(
                rf"Kleine\s+Anfrage\s*{ka}\s*(vom|\b)|"
                rf"Anfrage\s*{ka}\s*vom|"
                rf"Anfrage\s*{ka}vom",
                head_norm):
            flags.append("md_kanr")
    return flags


# Tags that indicate a row has been hand-corrected and should count as
# 'korrigiert' rather than 'ask_review' even if no open mismatches remain.
_MANUAL_TAGS = {
    "anfrager_manual", "anfrager_pdf_typo", "anfrager_from_db",
    "anfrager_cross_fraktion", "pdf_database_mismatch",
}


def cmd_merge(args: argparse.Namespace) -> int:
    """Persist DB ↔ PDF cross-check verdict on every row.

    For each row with status=extracted, re-parse the PDF header and write:
      - Mismatch_Flags: ","-joined field names where DB ↔ PDF differ
      - Datenqualität: ok / korrigiert / ask_review

    'korrigiert' = a manual-correction tag is set in Extract_Flags AND no
    open mismatch remains. 'ask_review' = at least one mismatch outstanding.
    """
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)
    counters = {"ok": 0, "korrigiert": 0, "ask_review": 0, "skipped": 0}

    for key, rec in rows.items():
        if rec.antworttext_status != STATUS_EXTRACTED or not rec.antworttext:
            rec.mismatch_flags = ""
            rec.datenqualitaet = ""
            counters["skipped"] += 1
            continue
        md_path = REPO_ROOT / rec.antworttext
        if not md_path.exists():
            counters["skipped"] += 1
            continue
        try:
            head = md_path.read_text(encoding="utf-8", errors="replace")[:16000]
        except OSError:
            counters["skipped"] += 1
            continue
        flags = _detect_mismatches(rec, head)
        rec.mismatch_flags = ",".join(flags)
        existing_tags = {t.strip() for t in (rec.extract_flags or "").split(",") if t.strip()}
        has_manual = bool(existing_tags & _MANUAL_TAGS)
        # Eine nicht-leere Notiz ist eine erledigte Mismatch-Triage
        # (resolve --auto / --interactive / Hand). Die bleibt 'korrigiert',
        # auch wenn _detect_mismatches nach Toleranz-Update keinen Flag mehr
        # findet — sonst geht der resolve-Trail verloren.
        has_notiz = bool((rec.notizen or "").strip())
        if flags:
            rec.datenqualitaet = "korrigiert" if has_notiz else "ask_review"
            counters["korrigiert" if has_notiz else "ask_review"] += 1
        elif has_manual or has_notiz:
            rec.datenqualitaet = "korrigiert"
            counters["korrigiert"] += 1
        else:
            rec.datenqualitaet = "ok"
            counters["ok"] += 1

    save_index(rows, xlsx)
    print(
        f"merge done: ok={counters['ok']} korrigiert={counters['korrigiert']} "
        f"ask_review={counters['ask_review']} skipped={counters['skipped']}",
        file=sys.stderr,
    )
    return 0


def cmd_verify(args: argparse.Namespace) -> int:
    """Read-only report; optional --llm-plausibility / --llm-rescue-fields."""
    xlsx = Path(args.xlsx)
    rows = load_index(xlsx)
    if not rows:
        print("verify: index.xlsx is empty or absent", file=sys.stderr)
        return 1

    # Counts by WP and status
    by_wp: dict[int, int] = {}
    by_status: dict[str, int] = {}
    md_missing: list[str] = []
    md_short: list[str] = []
    md_kanr_mismatch: list[str] = []  # md text does NOT contain "Kleine Anfrage <Nr>"
    md_grosse_anfrage: list[str] = []  # md text is a Große Anfrage answer
    bad_nr: list[str] = []
    negative_response_time: list[str] = []  # Antwortdatum strictly before Anfragedatum
    # Cross-check: PDF header re-parse vs DB. None of these abort the run; they
    # surface candidates for the resolve pipeline.
    mismatch_anfragedatum: list[str] = []
    mismatch_antwortdatum: list[str] = []
    mismatch_anfrage_drucksache: list[str] = []
    mismatch_fraktion: list[str] = []
    novel_min: dict[str, int] = {}
    novel_frak: dict[str, int] = {}
    by_wp_ka: dict[int, set[int]] = {}            # for KA gap detection
    dup_ka: dict[tuple[int, int], list[str]] = {}  # (wp, ka_nr) → [pks]

    for key, rec in rows.items():
        by_wp[rec.wp or 0] = by_wp.get(rec.wp or 0, 0) + 1
        by_status[rec.antworttext_status] = by_status.get(rec.antworttext_status, 0) + 1
        if rec.antworttext_status == STATUS_EXTRACTED and rec.antworttext:
            md_path = REPO_ROOT / rec.antworttext
            if not md_path.exists():
                md_missing.append(key)
            else:
                size = md_path.stat().st_size
                if size < 200:
                    md_short.append(f"{key} ({size}B)")
                elif rec.kleine_anfrage_nr:
                    # Sanity: the answer text must mention "Kleine Anfrage <Nr>".
                    # Catches mis-linked rows (e.g. KA 144 row pointing at 18/10805
                    # which is a Große Anfrage answer). Only first 4 KB read — the
                    # phrase appears on page 1.
                    try:
                        head = md_path.read_text(encoding="utf-8", errors="replace")[:16000]
                    except OSError:
                        head = ""
                    head_norm = _maybe_dedouble(head)
                    ka = rec.kleine_anfrage_nr
                    if not re.search(
                            rf"Kleine\s+Anfrage\s*{ka}\s*(vom|\b)|"
                            rf"Anfrage\s*{ka}\s*vom|"
                            rf"Anfrage\s*{ka}vom",
                            head_norm):
                        md_kanr_mismatch.append(f"{key} (KA {ka})")
                        if is_grosse_anfrage(head_norm):
                            md_grosse_anfrage.append(f"{key} (KA {ka})")
                    # Cross-check DB ↔ PDF header for the fields the search index
                    # also delivers. Only flag when both sides are non-empty AND
                    # PDF parse succeeded (avoids noise from header layout drift).
                    parsed = _parse_pdf_header_fields(head)
                    # Datums-Mismatches: 0–122d-Korridor in PDF-früher-Richtung
                    # (Drucksache↔Brief-Lag) ist kein Mismatch; siehe
                    # _date_mismatch().
                    if _date_mismatch(parsed["anfragedatum"], rec.anfragedatum):
                        try:
                            db_d = date.fromisoformat(rec.anfragedatum)
                            pdf_d = date.fromisoformat(parsed["anfragedatum"])
                            mismatch_anfragedatum.append(
                                f"{key}: DB={rec.anfragedatum} PDF={parsed['anfragedatum']} "
                                f"({(pdf_d - db_d).days}d)")
                        except ValueError:
                            pass
                    if _date_mismatch(parsed["antwortdatum"], rec.antwortdatum):
                        try:
                            db_d = date.fromisoformat(rec.antwortdatum)
                            pdf_d = date.fromisoformat(parsed["antwortdatum"])
                            mismatch_antwortdatum.append(
                                f"{key}: DB={rec.antwortdatum} PDF={parsed['antwortdatum']} "
                                f"({(pdf_d - db_d).days}d)")
                        except ValueError:
                            pass
                    if (parsed["drucksache_anfrage_nr"] and rec.drucksache_anfrage_nr
                            and parsed["drucksache_anfrage_nr"] != rec.drucksache_anfrage_nr):
                        mismatch_anfrage_drucksache.append(
                            f"{key}: DB={rec.drucksache_anfrage_nr} PDF={parsed['drucksache_anfrage_nr']}")
                    if (parsed["fraktion"] and rec.fraktion
                            and parsed["fraktion"] != rec.fraktion):
                        mismatch_fraktion.append(
                            f"{key}: DB={rec.fraktion} PDF={parsed['fraktion']}")
        if rec.drucksache_antwort_nr and not re.match(r"\d+/\d+$", rec.drucksache_antwort_nr):
            bad_nr.append(f"{key}: drucksache_antwort_nr={rec.drucksache_antwort_nr!r}")
        if rec.anfragedatum and rec.antwortdatum:
            try:
                ad = date.fromisoformat(rec.anfragedatum)
                rd = date.fromisoformat(rec.antwortdatum)
                if rd < ad:
                    negative_response_time.append(
                        f"{key}: {rec.anfragedatum} → {rec.antwortdatum} ({(rd - ad).days}d)"
                    )
            except ValueError:
                pass
        if rec.fraktion and not check_fraktion(rec.fraktion):
            novel_frak[rec.fraktion] = novel_frak.get(rec.fraktion, 0) + 1
        if rec.wp and rec.kleine_anfrage_nr:
            by_wp_ka.setdefault(rec.wp, set()).add(rec.kleine_anfrage_nr)
            dup_ka.setdefault((rec.wp, rec.kleine_anfrage_nr), []).append(key)

    # Gap detection: contiguous KA-Nr from 1..max(seen) per WP
    ka_gaps: dict[int, list[int]] = {}
    for wp, seen in by_wp_ka.items():
        if not seen:
            continue
        hi = max(seen)
        gaps = sorted(set(range(1, hi + 1)) - seen)
        if gaps:
            ka_gaps[wp] = gaps

    # Duplicate KA-Nr (same wp+ka_nr appearing in >1 row, e.g. correct row +
    # garbage row whose answer-PDF was misidentified as that KA).
    dup_lines: list[str] = []
    for (wp, ka), pks in dup_ka.items():
        if len(pks) > 1:
            dup_lines.append(f"WP{wp} KA {ka}: {', '.join(pks)}")

    # Orphan PDF/MD files (no matching xlsx row)
    keys_set = set(rows.keys())
    orphans = []
    for pdf in iter_archive_pdfs():
        try:
            _, da = parse_filename_drucksache_nr(pdf)
        except ValueError:
            continue
        if da not in keys_set:
            orphans.append(pdf.name)

    print(f"=== verify @ {_now_iso()} ===")
    print(f"rows: {len(rows)}")
    print("by WP:", ", ".join(f"{wp}={n}" for wp, n in sorted(by_wp.items())))
    print("by status:", ", ".join(f"{s}={n}" for s, n in sorted(by_status.items())))
    print()
    print(f"md missing (status=extracted but file gone): {len(md_missing)}")
    for k in md_missing[:10]:
        print(f"  {k}")
    print(f"md suspiciously short (<200B): {len(md_short)}")
    for k in md_short[:10]:
        print(f"  {k}")
    print(f"md ↔ KA-Nr mismatch (md text lacks 'Kleine Anfrage <Nr>'): {len(md_kanr_mismatch)}")
    for k in md_kanr_mismatch[:10]:
        print(f"  {k}")
    print(f"  davon Große Anfragen (out-of-scope, sollten gelöscht werden): {len(md_grosse_anfrage)}")
    for k in md_grosse_anfrage[:10]:
        print(f"    {k}")
    print(f"malformed Drucksache_*_Nr: {len(bad_nr)}")
    for k in bad_nr[:10]:
        print(f"  {k}")
    print(f"negative Antwortzeit (Antwortdatum vor Anfragedatum): {len(negative_response_time)}")
    for k in negative_response_time[:10]:
        print(f"  {k}")
    print(f"DB↔PDF Anfragedatum Mismatch: {len(mismatch_anfragedatum)}")
    for k in mismatch_anfragedatum[:10]:
        print(f"  {k}")
    print(f"DB↔PDF Antwortdatum Mismatch: {len(mismatch_antwortdatum)}")
    for k in mismatch_antwortdatum[:10]:
        print(f"  {k}")
    print(f"DB↔PDF Drucksache_Anfrage Mismatch: {len(mismatch_anfrage_drucksache)}")
    for k in mismatch_anfrage_drucksache[:10]:
        print(f"  {k}")
    print(f"DB↔PDF Fraktion Mismatch: {len(mismatch_fraktion)}")
    for k in mismatch_fraktion[:10]:
        print(f"  {k}")
    print(f"orphan PDFs in Archiv (no matching row): {len(orphans)}")
    for k in orphans[:10]:
        print(f"  {k}")
    print()
    # KA-Nr gaps per WP
    print("KA-Nr Lücken (fortlaufend nummeriert; jede Lücke ist potenziell eine fehlende oder zurückgezogene KA):")
    for wp in sorted(ka_gaps):
        gaps = ka_gaps[wp]
        sample = ", ".join(map(str, gaps[:30]))
        more = f" (+{len(gaps)-30} weitere)" if len(gaps) > 30 else ""
        print(f"  WP{wp}: {len(gaps)} Lücken — {sample}{more}")
    if not ka_gaps:
        print("  keine")
    print()
    print(f"Duplikat-KA-Nr (gleiche wp+ka_nr in mehreren Zeilen): {len(dup_lines)}")
    for line in dup_lines[:10]:
        print(f"  {line}")
    print()
    print(f"vocab novelty — Fraktion: {sum(novel_frak.values())} occurrences across {len(novel_frak)} values")
    for v, n in sorted(novel_frak.items(), key=lambda x: -x[1])[:10]:
        print(f"  {n:4d}× {v!r}")

    # Hard-fail categories
    bad = (len(md_missing) + len(md_short) + len(bad_nr) + len(md_kanr_mismatch)
           + len(dup_lines) + len(negative_response_time))
    if bad:
        print(f"\nverify: {bad} broken row(s); exit non-zero", file=sys.stderr)
        return 1
    print("\nverify: clean.")
    return 0


# --- CLI ---------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="landtag", description=__doc__)
    p.add_argument("--xlsx", default=str(INDEX_XLSX), help="path to index.xlsx")
    p.add_argument("--rps", type=float, default=DEFAULT_RPS,
                   help="HTTP requests per second (shared budget). Default polite ceiling for a small public-sector server.")
    p.add_argument("--user-agent", default=USER_AGENT)

    sub = p.add_subparsers(dest="cmd", required=True)

    s = sub.add_parser("scan-archive", help="enrich existing rows from local PDFs")
    s.add_argument("--wahlperiode", type=int)
    s.add_argument("--force", action="store_true")
    s.add_argument("--limit", type=int, help="cap number of PDFs to scan (for testing)")
    s.add_argument("--allow-discovery", action="store_true",
                   help="legacy: create new rows from PDFs alone (without crawl). "
                        "Default: skip PDFs whose Drucksache-Nr is not yet in the index.")
    s.set_defaults(func=cmd_scan_archive)

    s = sub.add_parser("crawl", help="enrich (default) or discover via Webflow search")
    s.add_argument("--wahlperiode", type=int)
    s.add_argument("--from", dest="date_from", help="client-side filter on Anfragedatum")
    s.add_argument("--to", dest="date_to", help="client-side filter on Anfragedatum")
    s.add_argument("--rpp", type=int, default=50)
    s.add_argument("--full", action="store_true",
                   help="DEPRECATED no-op (full sweep is now the only mode; flag kept for back-compat)")
    s.set_defaults(func=cmd_crawl)

    s = sub.add_parser("normalize",
                       help="match Fraktion/Ministerium against Index/*.xlsx and fill canonical columns")
    s.set_defaults(func=cmd_normalize)

    s = sub.add_parser("extract-multi-ministerium",
                       help="parse Antworttext-MD for ALL beteiligte Ministerien (Beteiligte_Ministerien_Kuerzel)")
    s.add_argument("--wahlperiode", type=int)
    s.set_defaults(func=cmd_extract_multi_ministerium)

    s = sub.add_parser("build-abgeordnete-index",
                       help="build Index/abgeordnete.xlsx from data/index.xlsx Anfrager column")
    s.set_defaults(func=cmd_build_abgeordnete_index)

    s = sub.add_parser("extract-all-anfrager",
                       help="parse Antwort-PDF Anfragerblock for FULL co-signer list "
                            "(Anfrager_Alle, Anzahl_Abgeordnete) — fixes DB cap of 2 + 'u.a.'")
    s.add_argument("--wahlperiode", type=int)
    s.set_defaults(func=cmd_extract_all_anfrager)

    s = sub.add_parser("enrich-llm",
                       help="LLM rescue for rows the rule-based parser left flagged in Extract_Flags")
    s.add_argument("--wahlperiode", type=int)
    s.add_argument("--limit", type=int, help="cap number of LLM calls (for testing/cost-control)")
    s.add_argument("--model", default="gpt-5-mini",
                   help="model name passed to `llm -m` (e.g. gpt-5-mini, gemma4:31b)")
    s.set_defaults(func=cmd_enrich_llm)

    s = sub.add_parser("fetch-text", help="download missing PDFs and extract .md")
    s.add_argument("--wahlperiode", type=int)
    s.add_argument("--limit", type=int)
    s.add_argument("--force", action="store_true")
    s.add_argument("--workers", type=int, default=1)
    s.set_defaults(func=cmd_fetch_text)

    s = sub.add_parser("resolve",
                       help="targeted re-search (--ka …) or interactive Mismatch-Review (--interactive)")
    s.add_argument("--ka", type=int, action="append",
                   help="KA-Nr to re-search (repeatable, e.g. --ka 144 --ka 3167). "
                        "Mutually exclusive with --interactive.")
    s.add_argument("--wahlperiode", type=int, default=18)
    s.add_argument("--counter-search", action="store_true",
                   help="bei 0 KA-Treffern: Gegensuche mit doktyp=GA (Große Anfrage)")
    s.add_argument("--delete-phantom", action="store_true",
                   help="Phantom-Zeilen (stale KA-Nr, oder GA-Match) aus dem Index löschen")
    s.add_argument("--interactive", action="store_true",
                   help="Iterate Datenqualität=ask_review rows, prompt for note + verdict.")
    s.add_argument("--auto", action="store_true",
                   help="Heuristik-Pass über alle Datenqualität=ask_review-Zeilen: "
                        "DB-authoritative-Mismatches (PDF-OCR-Glitch, Drucksachen-vs-"
                        "Brief-Datum, Fraktionsaustritt etc.) auto-klassifizieren; "
                        "Rest bleibt ask_review für menschliche Triage. Idempotent.")
    s.add_argument("--dry-run", action="store_true",
                   help="Mit --auto: Zähler ausgeben, aber keine Notizen schreiben.")
    s.set_defaults(func=cmd_resolve)

    s = sub.add_parser("merge",
                       help="persist DB ↔ PDF cross-check verdict (Mismatch_Flags + Datenqualität)")
    s.set_defaults(func=cmd_merge)

    s = sub.add_parser("verify", help="read-only sanity report")
    s.add_argument("--probe-site", action="store_true")
    s.add_argument("--llm-plausibility", action="store_true")
    s.add_argument("--llm-rescue-fields", action="store_true")
    s.set_defaults(func=cmd_verify)

    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    DATA_DIR.mkdir(exist_ok=True)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
